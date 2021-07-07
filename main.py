'''
[Author]

Developer:          Zaahier Adams
Github:             https://github.com/ZaahierAdams

Last updated:       29/02/2020

'''



'''

[About]

This is a data scrubber for PSETA NSD QSDR xlsx Workbooks. 
It provides an intuitive interface which enables the user to interact with the application.
Moreover, the application creates and saves reports on errors resolved as well 
as residual unresolved issues.

For detailed information on this application – see User Guide in Documents 


[Acknowledgements]
Application's icon created by user freepik from flaticon.com

'''


version_string  = '1.00'
last_updated    = '29/02/2020'



from tkinter import (Tk, 
                     Menu, 
                     Frame, 
                     Toplevel, 
                     Label, 
                     PhotoImage, 
                     Text, 
                     Button, Radiobutton, 
                     Scrollbar, 
                     mainloop,
                     TclError
                     )

from tkinter import (W, X, LEFT, BOTH, TOP, SOLID, CHAR, StringVar, INSERT, END)

from tkinter import messagebox
from tkinter import filedialog 

import matplotlib
matplotlib.use("Agg")
from matplotlib import pyplot as plt
from matplotlib.backends.backend_tkagg import (
    FigureCanvasTkAgg, NavigationToolbar2Tk)
#from matplotlib.backend_bases import key_press_handler
from matplotlib.figure import Figure



# Main Function 
# =============================================================================
# =============================================================================
# =============================================================================
# =============================================================================
# =============================================================================

from os import path, mkdir, chdir, startfile, getcwd
from glob import glob
from pandas import read_excel, ExcelWriter, isnull
from time import time 
from re import sub
from openpyxl import load_workbook
#import xlsxwriter 
from datetime import date, datetime
#from matplotlib import cm   # for pie chart
from numpy import array 
from xlrd.biffh import XLRDError

import Texts
from XLSMVBA import XLSM_Create, XLSM_Run
from Clear_Directory import Clear_Directory, Clear_file_type
from Move import Move_file, Copy_file
from XLSM_to_XLSX import XLSM_to_XLSX
from Combine_SimpleXLSX import Combine_XLSX
from Dirt_Report import Dirt_Report


#import os
#import glob
#import pandas as pd
#import time 
#import re
#from openpyxl import load_workbook
##import xlsxwriter 
#from datetime import date, datetime
#from matplotlib import pyplot as plt
#from matplotlib import cm 
#import numpy as np
#
#import Texts
#import XLSMVBA
#import Clear_Directory
#import Move
#import XLSM_to_XLSX
#import Combine_SimpleXLSX
#import Dirt_Report


#from System import Texts
#from System import XLSMVBA
#from System import Clear_Directory
#from System import Move
#from System import XLSM_to_XLSX
#from System import Combine_SimpleXLSX
#from System import Dirt_Report





abspath =  path.abspath(__file__)
dname =  path.dirname(abspath)

#print('_____________________________________________________\n')
#print('1',os.getcwd())
#print('_____________________________________________________\n\n')


def main(radio_outformat, radio_graphic, radio_graphic2, win_w, win_h):
    
#    print('_____________________________________________________\n')
#    print('2',os.getcwd())
#    print('_____________________________________________________\n\n')

    '''

    Main Function
    > Iterates through all Excel docs in Test dir
    > Iterates through each sheet
    > Sends sheets (DF) to be scrubbed in Verify_Write()
    > Saves number of errors fixed and unresolved
    > Creates graphs of errors 
    
    '''
    

    
    print(radio_outformat)
    print(radio_graphic)
    print(radio_graphic2)
    
     
    work_done = True
    
    
    ## File type
    file_type       = '*.xlsx'
    
    ## 
#    file_property_1 = 'COMBINED QMR'
#    ## Sheet name should contain:
#    SN_contain_1    = ['O-COME', '4.3.2'] 
#    SN_contain_2    = ['all training employed 18.1' , 
#        
#                       'all training unemployed 18.2']
    
    ## Added 13/01/2020
    ## list of sheets to ignore - crude fix  
    SN_contain_3    = ['STATS']
    
    '''
    Databases
    
        DB  - Persal
        DB2 - Training Interventions
        
    '''
    DB_dirname        = 'DB'
    
    ## Persal DB file details
    
    #DB_fname        = '(Lighter) INU320181213XX2019-Dir People Development-Filled-Excl Sessions.xls'
    #DB_sheet        = 'XX2019-20181213-Filled & Intern'
    
    DB_fname        = 'Persal_DB.xls'
    DB_sheet        = 'Persal' 
    
    DB_index_col    = 'ID Number'
    DB_date_col     = 'Today'
    
    
    ## Training Intervention DB file details
    DB_TI_fname       = 'Training_DB.xlsx'
    DB_TI_sheet       = 'Training'
    
    DB_TI_index_col   = 'Intervention'
    
    
    
    ## Columns to check for duplicates
    Dup_Check_1     = 'PERSAL NUMBER'
    Dup_Check_2     = 'NAME OF THE LEARNER' #Note !!! "Names" for QMR !!!
    Dup_Check_3     = 'SURNAME OF THE LEARNER'
    Dup_Check_4     = 'AGE'
    
    
## Unused:
##  Occ cat Labour          [5]
##  Address3                [6]
##  Job Title on Post       [9]

## Update column header for new persal doc 

##                        Persal                  district (all?)                     combined (all?)
    cols_verify     = [
                        ['FIRST NAMES',          Dup_Check_2,                        'NAMES OF THE LEARNER'     ],
                        ['Surname',              Dup_Check_3,                        'SURNAME OF THE LEARNER'   ],
                        ['Persalno',            'PERSAL NUMBER'                                                 ],
                        ['SL-Person',           'SALARY LEVEL'                                                  ],
                        ['ID Number',           'ID NUMBER OF THE LEARNER',         'ID NUMBER OF THE LEARNER'  ],
                        ['Occ Cat Labour',      'OCCUPATIONAL LEVEL'                                            ],
                        ['ADDRESS 3',           'SPECIFY LEARNER RESIDENTIAL  AREA'                             ],
                        ['Race',                'RACE'                                                          ],
                        ['Disabled',            'DISABILITY (YES/NO)',              'DISABILITY'                ],
                        ['Job Title on Rank',   'JOB TITLE'                                                     ]
                        
                        ]
                        
    #'FIRST NAMES'
    #'Surname'
    #'Persalno'
    #'SL-Person'
    #'ID Number'
    #'Occ Cat Labour'
    #'ADDRESS 3'
    #'Race   '
    #'Disabled'
    #'Job Title on Post'
    
##                        Training DB                      QMR                         Training DB 
    cols_verify_TI   = [
                       ['Intervention',         'TRAINING INTERVENTION'                                         ],
                       ['Type',                 'TYPE OF LEARNING PROGRAMME'                                    ],
                       ['NQF Level',            'NQF LEVEL'                                                     ],
                       ['Duration',             'DURATION OF PROGRAMME',            'Unit Of Measure'           ],
                       ['Provider / Source',    'NAME OF THE TRAINING PROVIDER'                                 ]
                       
                       ]
    
    
    DB_col_keep     = [cols_verify[0][0], 
                       cols_verify[1][0], 
                       cols_verify[2][0],
                       cols_verify[3][0],
                       cols_verify[4][0],
                       cols_verify[5][0],
                       cols_verify[6][0],
                       cols_verify[7][0],
                       cols_verify[8][0],
                       cols_verify[9][0],
                       DB_date_col] 
    

    DB_col_keep_TI   = [cols_verify_TI[0][0], 
                        cols_verify_TI[1][0], 
                        cols_verify_TI[2][0],
                        cols_verify_TI[3][0], cols_verify_TI[3][2],
                        cols_verify_TI[4][0],
                       ] 
               
    ## Index 
    file_index_col  = 'ID NUMBER OF THE LEARNER'
      
#    Sheet_header   = 0
   
    
    
    
    
    '''
    Create Directories 
    '''
    ## File storage names
    folder_name_1   = 'Input files'
    folder_name_2   = 'Output files'
    folder_name_6   = 'Result Diagrams'
    
    folder_name_3   = 'Separate Sheets'
    folder_name_4   = 'XLSM'
    folder_name_5   = 'Formatted Workbooks'
    folder_name_7   = 'Simple Workbooks'
    folder_name_8   = 'Result Reports'
    
    dir_list_1 = [folder_name_1,
                folder_name_2, 
                folder_name_6,
                folder_name_8]

    dir_list_2 = [folder_name_3,
                folder_name_4, 
                folder_name_5,
                folder_name_7]
    
    for dirname1 in dir_list_1:
        try:
            mkdir(dirname1)
        except:
            pass
    
    for dirname2 in dir_list_2:
        try:
            mkdir(folder_name_2 + '\\' + dirname2)
        except:
            pass
        
#
#    ## directory for Test excel files
#    try:
#        os.mkdir(folder_name_1)
#    except:
#        pass
#    
#    ## directory for New excel files
#    try:
#        os.mkdir(folder_name_2)
#    except:
#        pass
#    
#    ## directory for seperated worksheets
#    try:
#        os.mkdir(folder_name_2+'\\'+folder_name_3)
#    except:
#        pass
#    
#    ## directory for .xlsm files
#    try:
#        os.mkdir(folder_name_2+'\\'+folder_name_4)
#    except:
#        pass
#    
#    ## directory for final .xlsx files
#    try:
#        os.mkdir(folder_name_2+'\\'+folder_name_5)
#    except:
#        pass
#    
#    ## directory for matplotlib diagrams
#    try:
#        os.mkdir(folder_name_6)
#    except:
#        pass
#
#    ## directory for simple .xlsx files
#    try:
#        os.mkdir(folder_name_2+'\\'+folder_name_7)
#    except:
#        pass
#
#    ## directory for Dirt Reports
#    try:
#        os.mkdir(folder_name_8)
#    except:
#        pass
    
    
    "Iterate through all Excel files in specified dir."
    file_list = []
    #file_type_len = len(file_type)
    parent_Dir  = getcwd()
    
    ## Dir for DB 
    DBDir       = (parent_Dir 
                   + '\\''' 
                   + DB_dirname 
                   + '\\''' 
                   + DB_fname)
    
    ## Dir for test xlsx files 
    file_Dir_1  = parent_Dir + '\\''' + folder_name_1
    chdir(file_Dir_1)

    ## Feedback notification #1 
    feedback.delete('1.0', END)
#    feedback.insert(INSERT, "⯈ Loading reference Persal DB...")
#    feedback.insert(END, '\n\n')    
    
    ## Read databases
    DB_start    = time()
    try:
        df_db       = read_excel(DBDir, 
                                sheet_name = DB_sheet,
                                index_col = DB_index_col,
                                usecols =   DB_col_keep)
        feedback.configure(bg = '#ffffff')
    except FileNotFoundError:
        
        Critical_Error('⯈ Persal Database or DB directory is missing', 
                       '\n⯈ Ensure that Persal Database file is named as \'Persal_DB\'' )
        
                           
        chdir(dname)
        return 
    except XLRDError:
        
        Critical_Error('⯈ No sheet named \'Persal\' in Persal_DB', 
                       '\n⯈ Rename appropriate sheet in \'Persal_DB\' to \'Persal\''  )
        
        chdir(dname)
        return
    except ValueError as val_errmsg:
        Critical_Error_2('⯈ The Persal database is missing a critical column !')
        
        feedback.insert(INSERT, 'More information:'
                        + '\n'
                        + str(val_errmsg))
        feedback.insert(END,'\n\n')  
        feedback.tag_add("description_3", "6.0", "6.17")

        chdir(dname)
        return 
        
    
    ## Checks on ref DB
    df_db = df_db.loc[~df_db.index.isna()]
    df_db = df_db.loc[~df_db.index.duplicated(keep='first')]                    
    DB_end      = time()
    DB_elapse   = str(round(DB_end - DB_start))
    DB_elapse_s = ('⯈ Took {} seconds to load Persal DB'.format(DB_elapse))                      # TEMP Notification 
    
    
    ## Feedback notification #2
    feedback.insert(INSERT, DB_elapse_s)
    feedback.insert(END, '\n\n')    

    ## Persal DB properties
    DB_shape    = df_db.shape 
    DB_info     = df_db.info()
    DB_info2    = df_db.info(memory_usage='deep')
    DB_info3    = df_db.memory_usage(deep=True)
    DB_info4    = df_db.memory_usage(deep=True).sum()
    
    
    ## Date and age of Persal DB
    DB_date     = str(list(set(df_db[DB_date_col].tolist()))[0])
    DB_date_YY  = DB_date[0:4]
    DB_date_MM  = DB_date[4:6]
    DB_date_DD  = DB_date[6:]
    DB_date_full= DB_date_DD+'/'+DB_date_MM+'/'+DB_date_YY
    
    DB_Age = Calc_Age(date(int(DB_date_YY), int(DB_date_MM), int(DB_date_DD)))
    if DB_Age >= 1:
        qual_ageDB = '(Persal DB needs updating!)'
    else:
        qual_ageDB = ''
    
   
    print('XXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXX')
    print('Date of Persal DB:', DB_date_full,qual_ageDB)                        # TEMP Notification 
    print()
    print(DB_shape)
    print(DB_info2)
    print(DB_info3)
    print('XXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXX')
    ##   int. > str. (stace efficiency)
    
    ## Feedback notification #3
    feedback.insert(INSERT, '⯈ Date of Persal DB: ' + DB_date_full + ' ' +qual_ageDB)
    feedback.insert(END, '\n\n')  
    
    
    
    
    
  
    ' Read TI DB '
    
    ## Dir for TI DB 
    DB_TI_Dir       = (parent_Dir 
                       + '\\''' 
                       + DB_dirname 
                       + '\\''' 
                       + DB_TI_fname)
    
#    ## Dir for test xlsx files 
#    chdir(file_Dir_1)

#    feedback.insert(INSERT, "⯈ Loading reference Training Intervention DB...")
#    feedback.insert(END, '\n\n')    
    
    ## Read databases
#    DB_start_TI   = time()
    try:
        df_db_TI      = read_excel(DB_TI_Dir, 
                                    sheet_name = DB_TI_sheet,
                                    index_col  = DB_TI_index_col,
                                    usecols    = DB_col_keep_TI)
        feedback.configure(bg = '#ffffff')
    except FileNotFoundError:
        feedback.delete('1.0', END)
        Critical_Error('⯈ Training Interventions Database is missing!', 
                      '\n⯈ Ensure that Training Interventions file is named as \'Training_DB\'' )
        chdir(dname)
        return
    except XLRDError:
        feedback.delete('1.0', END)
        Critical_Error('⯈ No sheet named \'Training\' in Training_DB', 
                      '\n⯈ Rename appropriate sheet in \'Training_DB\' to \'Training\'' )
        chdir(dname)
        return
    except ValueError as val_errmsg:
        feedback.delete('1.0', END)
        Critical_Error_2('⯈ The Training_DB is missing a critical column !')
        feedback.insert(INSERT, 'More information:' + '\n' + str(val_errmsg))
        feedback.insert(END,'\n\n')  
        feedback.tag_add("description_3", "6.0", "6.17")
        chdir(dname)
        return 
        
#    print(df_db_TI)
    
    '       Disabled check for duplicates/ NaN - needs fixing! '
    ## Checks on TI ref DB
#    df_db_TI = df_db.loc[~df_db_TI.index.isna()]
#    df_db_TI = df_db.loc[~df_db_TI.index.duplicated(keep='first')]  
                  
#    DB_end_TI      = time()
#    DB_elapse_TI   = str(round(DB_end_TI - DB_start_TI))
#    DB_elapse_TI_s = ('⯈ Took {} seconds to load Training Intervention DB'.format(DB_elapse_TI))                     
#    feedback.insert(INSERT, DB_elapse_TI_s)
#    feedback.insert(END, '\n\n')
    

    
    
    
     
    
    ## Variables: All workbook's blunders
    all_duplicate_rows      = 0
    all_no_err_ID           = 0
    all_unverify_ID         = 0
    all_mis_age             = 0
    all_mis_gender          = 0
    all_mis_citizen         = 0 
    all_mis_youth           = 0

    all_mis_name            = 0
    all_mis_sname           = 0
    all_mis_persal          = 0
    all_mis_SLvL            = 0
    all_mis_race            = 0
    all_mis_disab           = 0
    
    all_mis_occlvl          = 0
    all_mis_jobtit          = 0
    all_mis_res             = 0
    
    all_unveri_TI           = 0
    all_intervention        = 0
    all_TItype              = 0
    all_NQF                 = 0
    all_duration            = 0
    all_provider            = 0
    
    all_res_issues          = 0 

    counter_book            = 0
    counter_all_sheets      = 0
    
    ## list of all unverified ID's 
    all_khangela = []
    
    
    ## Clear txt temp reports
    Clear_file_type(parent_Dir, folder_name_1, 'txt')
    

    "Iterate through all files in Test Dir."
    for file in glob(file_type):
        
        file_list.append(file)
#        print(file)
#        print(file_list)
        
        ## Feedback notification #4
        feedback.insert(INSERT, '⯈ Scanning workbook: ' + file)
        feedback.insert(END, '\n\n')  

        df0 = read_excel(file,sheet_name = None, header = None)
        
    
        ## Variables: workbook's blunders
        file_duplicate_rows     = 0
        file_no_err_ID          = 0
        file_no_verify_ID       = 0
        file_mis_age            = 0
        file_mis_gender         = 0
        file_mis_citizen        = 0 
        file_mis_youth          = 0
        
        file_mis_name           = 0
        file_mis_sname          = 0
        file_mis_persal         = 0
        file_mis_SLvL           = 0
        file_mis_race           = 0
        file_mis_disab          = 0
        
        file_mis_occlvl         = 0
        file_mis_jobtit         = 0
        file_mis_res            = 0
        
        file_unveri_TI          = 0
        file_intervention       = 0
        file_TItype             = 0
        file_NQF                = 0
        file_duration           = 0
        file_provider           = 0
        
        file_res_issues         = 0 
        
        
        #======================================================================
        "start: GENERAL soln."
        
#        print('=====================================\nFILE =',file)
#        print()
        
        ## Dict that will contain sheet names that contain relevant header names,
        ##   and the rows #s with the valid headers
        dict_treat = {}
        list_treat = []
                
        ## List of all sheets
        all_sheets = []
        
        ## List of sheets to khangela
        khangela   = [] #resets for every book
        
        for name, df in df0.items():
            
#            ## Feedback notification #5
#            feedback.insert(INSERT, '\t⯈ Scanning worksheet: ' + name)
#            feedback.insert(END, '\n\n')  
            
            all_sheets.append(name)
            
            ## Variables to store row number of header columns
            ## Allows for 2 rows of headers columns
            rc1 = 0
            rc2 = 0
            
            two = 1
            
            ## Name of sheet
            sheetname = ''
            
            "fix 13/01/2020"
            "Crude fix - ignore sheets in list"
            if name not in SN_contain_3:
                
                for index, row in df.iterrows():
                    
                    ## Converts row to list
                    JJ = list(row)
                    
                    ## Check if any cells in row is an accepted column header
                    result = [e for e in JJ for i in cols_verify if e in i]
                    
                    ## Row count 
                    rc1 += 1
                    
                    ## If row contains atleast (1) accepted column header
                    if len(result) != 0:
                        
                        dict_treat[name] = rc1
                        if name in list_treat:
                            pass
                        else:
                            list_treat.append(name)
    
                        
                        ## If two columns within a sheet contains valid headers then
                        ## the for loop (above) will output two outputs belonging to the same sheet
                        ## This test is to find if there are two outputs belonging to the same sheet 
                        if two == 1:
                            sheetname = name     
                            ## Save first row number containing valid header columns as rc2
                            rc2 = rc1
                            two += 1
                            
                        elif two == 2:
                            ## This condition is only satisfied 
                            if sheetname == name:
                                "combine two columns here"
                                #print('yes')
    
                                del dict_treat[name]
                                
                                ## Places header columns in a list
                                ## (At this point, they are irrelevant sequential numbers)
                                ## The goal is to get the length of list of columns
                                columns = list(df)   
                                ## first row of headers will assume positions [0 to (n-6)]
                                len_1st_lvl_col = len(columns)-6
                                len_2nd_lvl_col = len(columns)
                                
                                ## "Merge" the two header columns
                                iter1 = 0
                                for i in columns:
                                    if iter1 >= len_1st_lvl_col:
                                        if iter1 == len_2nd_lvl_col:
                                            break
                                        else:
                                            iter1+=1
                                    else:
                                        df.at[rc1-1, i] = df[i][rc2-1]
                                        iter1+=1
                                df.columns = df.iloc[rc1-1]

                                
                                ## Drop "blank" columns above column headers
                                iter2 = 0
                                while iter2 < rc1:
                                    df = df.drop([iter2,iter2])
                                    iter2 += 1
                                
                                ## /NW fix - remove NaN columns
                                df = df.loc[:,df.columns.notnull()]
                                
                                #print(df.columns)
                                #print(df)
                                #print()
                                
                                "Set Index"
                                df, dup_rows, IDError = DF_form(name, df, 
                                                           file_index_col,  
                                                           Dup_Check_3,
                                                           Dup_Check_4)
                                
    #                            print(name, df)
    #                            print(name, df.columns)
    #                            print(name, df)
    #                            print()
    #                            print()
    #                            print()
                                
    
                                if IDError is False:
                                    "Verify and write DF"
                                    "Verify_Write"
                                    
                                                                                                                                                                                                                                                            
                                    duplicate_rows, no_err_ID, no_unverify_ID, mis_age, mis_gender, mis_citizen, mis_youth, Glob_mis_name, Glob_mis_sname, Glob_mis_persal, Glob_mis_SLvL, Glob_mis_race, Glob_mis_disab, Glob_mis_occlvl, Glob_mis_jobtit, Glob_mis_res, Glob_unveri_TI, Glob_intervention, Glob_TItype, Glob_NQF, Glob_duration, Glob_provider, Glob_res_issues  = Verify_Write(df_db, df_db_TI,
                                                         df, 
                                                         file, 
                                                         name, 
                                                         parent_Dir, 
                                                         folder_name_2,
                                                         folder_name_3, 
                                                         folder_name_8,
                                                         file_type,
                                                         dup_rows,
                                                         Dup_Check_1,
                                                         Dup_Check_2,
                                                         Dup_Check_3,
                                                         cols_verify
                                                         )  
                                    
                                    "Cleanse file stats"
                                    file_duplicate_rows     += duplicate_rows
                                    file_no_err_ID          += no_err_ID
                                    file_no_verify_ID       += no_unverify_ID
                                    file_mis_age            += mis_age
                                    file_mis_gender         += mis_gender
                                    file_mis_citizen        += mis_citizen
                                    file_mis_youth          += mis_youth
                                    
                                    file_mis_name           += Glob_mis_name
                                    file_mis_sname          += Glob_mis_sname
                                    file_mis_persal         += Glob_mis_persal
                                    file_mis_SLvL           += Glob_mis_SLvL
                                    file_mis_race           += Glob_mis_race
                                    file_mis_disab          += Glob_mis_disab
                                    
                                    file_mis_occlvl         += Glob_mis_occlvl
                                    file_mis_jobtit         += Glob_mis_jobtit
                                    file_mis_res            += Glob_mis_res
                                    
                                    file_unveri_TI          += Glob_unveri_TI 
                                    file_intervention       += Glob_intervention
                                    file_TItype             += Glob_TItype
                                    file_NQF                += Glob_NQF
                                    file_duration           += Glob_duration
                                    file_provider           += Glob_provider
                                    
                                    file_res_issues         += Glob_res_issues
    
    
                                    
                                    
                                    
                                    
                                    ## Note sheet names that contain unveried (Persal) ID Numbers
                                    if no_unverify_ID != 0:
                                        khangela.append(name)
                                    else:
                                        pass
    
                                    
                                    ## Does not count residual sheets:
                                    counter_all_sheets      += 1
                                else:
                                    ## Bypass sheet if IDError is True
                                    continue 
                                
                                
                            else: # surname = 
                                pass
                        else: # two = 
                            pass
                    else: # result == 0 (row)
                        pass
            else:
                pass
        
#        print(dict_treat)



        ## Soln to files /w one row of column headers:
        for name, df in df0.items():
            if name in dict_treat.keys():
                #print(name,'\t',dict_treat[name])
                
                rc1 = dict_treat[name]
                
                df.columns = df.iloc[rc1-1]
                 
                iter2 = 0
                while iter2 < rc1:
                    df = df.drop([iter2,iter2])
                    iter2 += 1
                
#                print(df)
                    
                ## /NW fix - remove NaN columns
                df = df.loc[:,df.columns.notnull()]
                    
                "Set Index"
                df, dup_rows, IDError = DF_form(name, df, 
                                       file_index_col,  
                                       Dup_Check_3,
                                       Dup_Check_4)
                
                if IDError is False:
                    "Verify and write DF"
                    duplicate_rows, no_err_ID, no_unverify_ID, mis_age, mis_gender, mis_citizen, mis_youth, Glob_mis_name, Glob_mis_sname, Glob_mis_persal, Glob_mis_SLvL, Glob_mis_race, Glob_mis_disab, Glob_mis_occlvl, Glob_mis_jobtit, Glob_mis_res, Glob_unveri_TI, Glob_intervention, Glob_TItype, Glob_NQF, Glob_duration, Glob_provider, Glob_res_issues = Verify_Write(df_db, df_db_TI,
                                             df, 
                                             file, 
                                             name, 
                                             parent_Dir, 
                                             folder_name_2,
                                             folder_name_3, 
                                             folder_name_8,
                                             file_type,
                                             dup_rows,
                                             Dup_Check_1,
                                             Dup_Check_2,
                                             Dup_Check_3,
                                             cols_verify
                                             )  
                    
                    "Cleanse file stats"
                    file_duplicate_rows     += duplicate_rows
                    file_no_err_ID          += no_err_ID
                    file_no_verify_ID       += no_unverify_ID
                    file_mis_age            += mis_age
                    file_mis_gender         += mis_gender
                    file_mis_citizen        += mis_citizen
                    file_mis_youth          += mis_youth
                    
                    file_mis_name           += Glob_mis_name
                    file_mis_sname          += Glob_mis_sname
                    file_mis_persal         += Glob_mis_persal
                    file_mis_SLvL           += Glob_mis_SLvL
                    file_mis_race           += Glob_mis_race
                    file_mis_disab          += Glob_mis_disab
                    
                    file_mis_occlvl         += Glob_mis_occlvl
                    file_mis_jobtit         += Glob_mis_jobtit
                    file_mis_res            += Glob_mis_res
                    
                    file_unveri_TI          += Glob_unveri_TI 
                    file_intervention       += Glob_intervention
                    file_TItype             += Glob_TItype
                    file_NQF                += Glob_NQF
                    file_duration           += Glob_duration
                    file_provider           += Glob_provider
                    
                    file_res_issues         += Glob_res_issues
                    
    
                    
                    if no_unverify_ID != 0:
                        khangela.append(name)
                    else:
                        pass
                    
                    counter_all_sheets      += 1
                                   
    #                print(name, df)
    #                print()
    #                print()
                else:
                    continue
                
            else:
                pass
        
        

        '''
        Solution for residual pages 
        
        '''
        
        ## list of residual sheets
        residual_sheets = [k for k in all_sheets if k not in list_treat]
        
        '''
            Method 1:
                Rough & dirty fix - Unformated pages
        '''
#        for name, df in df0.items():
#            if name in residual_sheets:
                  
#                file_Dir_2 = (parent_Dir 
#                  + '\\''' 
#                  + folder_name_2
#                  + '\\''' 
#                  + folder_name_3 
#                  + '\\''' 
#                  + name.replace(file_type.replace('*',''),'')
#                  + file_type.replace('*','')
#                  )    
#                df.to_excel(file_Dir_2, sheet_name = name)
#            else:
#                pass
                
        '''
            Method 2:
                openpyxl formatted
        '''
        for res_s in residual_sheets:
            wbX = load_workbook( file_Dir_1 + '\\' + file )
            sheets = wbX.sheetnames 
            
            for s in sheets:
                if s !=res_s:
                    sheet_name = wbX.get_sheet_by_name(s)
                    wbX.remove_sheet(sheet_name)
                else:
                    pass
        
            file_Dir_2 = (parent_Dir 
                  + '\\''' 
                  + folder_name_2
                  + '\\''' 
                  + folder_name_3 
                  + '\\''' 
                  + res_s
                  + file_type.replace('*','')
                  )  
        
            wbX.save(file_Dir_2)
            
            
                
        
        "end: GENERAL soln."
        #======================================================================
        
        #======================================================================
        "start: UNIQUE soln."
        "(artefact)"
#        
#        '''
#        Header format
#        '''
#        ## QMR format (hierarchy)
#        if file_property_1 in file:
#            
#            ## Iterate through each Excel Sheet
#            ## df    = dataframe for sheet
#            ## name  = name of sheet 
#            
#            counter_sheet       = 0
#            
#            for name, df in df0.items():
#                
#                ## Test to check if sheet name is valid 
#                valid_val = False
#                for test1 in SN_contain_1:
#                    if test1 in name:
#                        valid_val = True
#                        break
#                    else:
#                        pass
#                    
#                ## Worksheet name satisfies condition
#                if valid_val is True:
#                    
#                    ## Unify header hierarchy
#                    columns = list(df)   
#                    len_1st_lvl_col = len(columns)-6
#                    len_2nd_lvl_col = len(columns)
#                    iter1 = 0
#                
#                    for i in columns:
#                        if iter1 >= len_1st_lvl_col:
#                            if iter1 == len_2nd_lvl_col:
#                                break
#                            else:
#                                iter1+=1
#                        else:
#                            df.at[1,i] = df[i][0]
#                            iter1+=1
#                    
#                    ## select row as column Headers
#                    df.columns = df.iloc[1]             
#                    df = df.drop([0,1])
#                    
#
#                    df, dup_rows = DF_form(df, 
#                                           file_index_col, 
#                                           Dup_Check_3
#                                           )
#
#                        
#                    duplicate_rows, no_err_ID, mis_age, mis_gender, mis_citizen, mis_youth, Glob_mis_name, Glob_mis_sname, Glob_mis_persal, Glob_mis_SLvL, Glob_mis_race, Glob_mis_disab = Verify_Write(df_db,
#                                                                                                          df, 
#                                 file, 
#                                 name, 
#                                 parent_Dir, 
#                                 folder_name_2,
#                                 folder_name_3, 
#                                 file_type,
#                                 dup_rows,
#                                 Dup_Check_1,
#                                 Dup_Check_2,
#                                 Dup_Check_3,
#                                 cols_verify
#                                 )  
#                    
#                    file_duplicate_rows     += duplicate_rows
#                    file_no_err_ID          += no_err_ID
#                    file_mis_age            += mis_age
#                    file_mis_gender         += mis_gender
#                    file_mis_citizen        += mis_citizen
#                    file_mis_youth          += mis_youth
#                    
#                    counter_all_sheets      += 1
#                    
#                else:
#                    '''
#                    Solution for residual pages 
#                    
#                    rough & dirty fix !!! 
#                    
#                    Unformated pages 
#                    '''
#                    file_Dir_2 = (parent_Dir 
#                      + '\\''' 
#                      + folder_name_2
#                      + '\\''' 
#                      + folder_name_3 
#                      + '\\''' 
#                      + name.replace(file_type.replace('*',''),'')
#                      + file_type.replace('*','')
#                      )    
#                    df.to_excel(file_Dir_2, sheet_name = name)
#                    
#                
#                
#  
#        ## Other format
#        else:
#            for name, df in df0.items():
#                
#                if name in SN_contain_2: 
#                    
#                    ## Header size
#                    if name == SN_contain_2[0]:
#                        Sheet_header = 7
#                        
#                    elif name == SN_contain_2[1]:
#                        Sheet_header = 6
##                    else:
##                        pass
#                        
#                    file_list.append(file) #remove
#                    
##                    ## Dataframe
##                    df = pd.read_excel(file, 
##                                       sheet_name = name, 
##                                       skiprows = Sheet_header,
##                                       index_col = file_index_col)
#                    
#                    df.columns = df.iloc[Sheet_header - 1]  ## -1???
#                    
#                    iter2 = 0
#                    while iter2 < Sheet_header:
#                        ## index-row
#                        df = df.drop([iter2,iter2])
#                        iter2 += 1
#                    
#                    
#                     # REMOVE:
##                    ## select column as index
##                    df = df.set_index(file_index_col) 
##                    ## drop NaN values in index
##                    df = df.loc[~df.index.isna()]
##                    
##                    ## Rows with duplicate indices
##                    rows_all = df.shape[0] 
##                    ## drop duplicate indices 
##                    df = df.loc[~df.index.duplicated(keep='first')]
##                    ## Rows ~duplicate indices
##                    rows_after = df.shape[0]
##                    ## Number of duplicate rows
##                    dup_rows = rows_all - rows_after
#                        
#                        
#                    df, dup_rows = DF_form(df, 
#                                           file_index_col, 
#                                           Dup_Check_3
#                                           )
#                    
#                    duplicate_rows, no_err_ID, mis_age, mis_gender, mis_citizen, mis_youth, Glob_mis_name, Glob_mis_sname, Glob_mis_persal, Glob_mis_SLvL, Glob_mis_race, Glob_mis_disab = Verify_Write(df_db, df, 
#                                 file, 
#                                 name, 
#                                 parent_Dir, 
#                                 folder_name_2,
#                                 folder_name_3, 
#                                 file_type,
#                                 dup_rows,
#                                 Dup_Check_1,
#                                 Dup_Check_2,
#                                 Dup_Check_3,
#                                 cols_verify
#                                 )
#                    
#                    file_duplicate_rows     += duplicate_rows
#                    file_no_err_ID          += no_err_ID
#                    file_mis_age            += mis_age
#                    file_mis_gender         += mis_gender
#                    file_mis_citizen        += mis_citizen
#                    file_mis_youth          += mis_youth
#                    
#                    file_mis_name           += Glob_mis_name
#                    file_mis_sname          += Glob_mis_sname
#                    file_mis_persal         += Glob_mis_persal
#                    file_mis_SLvL           += Glob_mis_SLvL
#                    file_mis_race           += Glob_mis_race
#                    file_mis_disab          += Glob_mis_disab
#                    
#                    counter_all_sheets      += 1
#
#                
#                else:
#                    '''
#                    Missing Soln for residual pages !!! 
#                    '''
#                    pass
        "end: UNIQUE soln."
        #======================================================================



        ##  (TEMP) book Results 
        print( '---------------------------------------'                    )
        print( 'BOOK Duplicate rows:\t',        file_duplicate_rows         )
        print( 'BOOK Erroneous ID\'s:\t',       file_no_err_ID              )
        print( 'BOOK Unverified ID\'s:\t',      file_no_verify_ID           )
        print( 'BOOK Age blunders:\t',          file_mis_age                )
        print( 'BOOK Gendr blunders:\t',        file_mis_gender             )
        print( 'BOOK Citiz blunders:\t',        file_mis_citizen            )
        print( 'BOOK Youth blunders:\t',        file_mis_youth              )
        print(                                                              )
        print( 'BOOK Occ. Lvl blunders:\t',     file_mis_occlvl             )
        print( 'BOOK Job Title blunders:',      file_mis_jobtit             )
        print( 'BOOK Residence blunders:',      file_mis_res                )   
        print(                                                              )
        print( 'BOOK First Name:\t',            file_mis_name               )
        print( 'BOOK Surname:\t\t',             file_mis_sname              )
        print( 'BOOK Persal No.:\t',            file_mis_persal             )
        print( 'BOOK Salary Level:\t',          file_mis_SLvL               )
        print( 'BOOK Race:\t\t',                file_mis_race               )
        print( 'BOOK Disability:\t',            file_mis_disab              )
        print(                                                              )
        print( 'BOOK Unverified T.I.:\t',       file_unveri_TI              )
        print( 'BOOK Interventions:\t',         file_intervention           )
        print( 'BOOK Type:\t\t',                file_TItype                 )
        print( 'BOOK NQF:\t\t',                 file_NQF                    )
        print( 'BOOK duration:\t\t',            file_duration               )
        print( 'BOOK proivder:\t\t',            file_provider               )
        print(                                                              )
        print( 'BOOK Res. issues:\t',           file_res_issues             )
        print( '---------------------------------------'                    )
    
    
    
    
        file_n = file.replace(file_type.replace('*',''),'')
        
        
        all_duplicate_rows      += file_duplicate_rows
        all_no_err_ID           += file_no_err_ID
        all_unverify_ID         += file_no_verify_ID
        all_mis_age             += file_mis_age
        all_mis_gender          += file_mis_gender
        all_mis_citizen         += file_mis_citizen 
        all_mis_youth           += file_mis_youth
        

        all_mis_name            += file_mis_name
        all_mis_sname           += file_mis_sname
        all_mis_persal          += file_mis_persal
        all_mis_SLvL            += file_mis_SLvL
        all_mis_race            += file_mis_race
        all_mis_disab           += file_mis_disab
        
        all_mis_occlvl          += file_mis_occlvl
        all_mis_jobtit          += file_mis_jobtit
        all_mis_res             += file_mis_res
        
        all_unveri_TI           += file_unveri_TI
        all_intervention        += file_intervention
        all_TItype              += file_TItype
        all_NQF                 += file_NQF
        all_duration            += file_duration
        all_provider            += file_provider
        
        all_res_issues          += file_res_issues


        
        for jonga2 in khangela:
            all_khangela.append(file_n)

        counter_book           += 1
        
        Dirt_Report(parent_Dir, 
                    folder_name_8,
                    file_duplicate_rows,
                    file_no_err_ID,
                    file_no_verify_ID,
                    file_mis_age,
                    file_mis_gender,
                    file_mis_citizen,
                    file_mis_youth,
                    file_mis_name,
                    file_mis_sname,
                    file_mis_persal,
                    file_mis_SLvL,
                    file_mis_race,
                    file_mis_disab,
                    '-',
                    1,
                    file,
                    file_n,
                    2,
                    khangela,
                    file_mis_occlvl,
                    file_mis_jobtit,
                    file_mis_res,
                    file_unveri_TI,
                    file_intervention,
                    file_TItype,
                    file_NQF,
                    file_duration,
                    file_provider,
                    file_res_issues
                    )
                    # No sheet counter for each book !!! 


        '''
        Output products
        '''
        
        
        if radio_outformat == 'Formatted Workbook':
            '''
            VBA
            '''
            
            ## Create .xlsm file with macro
            XLSM_Create(parent_Dir, folder_name_2, file_n) 
            
            ## Run macro on .xlsm file
            VBA_fail = XLSM_Run(parent_Dir, folder_name_2, file_n)
            
            if VBA_fail is False:
                ## Empty sheets dir
                Clear_Directory(parent_Dir, folder_name_2, folder_name_3)
                
                ## Move .xlsm files to xlsm dir. 
                Move_file(parent_Dir, folder_name_2, folder_name_4, file_n)
                
                ## Convert .xlsm file to .xlsx 
                XLSM_to_XLSX(parent_Dir, folder_name_2, folder_name_4, folder_name_5, file_n)
                
                ## Open directory in Windows
    #            feedback.insert(INSERT, '⯈ Opening output directory')
    #            feedback.insert(END, '\n\n') 
                startfile( parent_Dir + '\\' + folder_name_2 + '\\' + folder_name_5 )
                
    #            print('_____________________________________________________\n')
    #            print('3',os.getcwd())
    #            print('_____________________________________________________\n\n')
            else:
                feedback.insert(INSERT, '⯈ ERROR! Failed to combine sheets!')
                feedback.insert(END, '\n\n') 
               

        elif radio_outformat == 'Separate Sheets':
            '''
            Seperate Sheets
            (Default option)
            '''
            
            ## Open directory in Windows
#            feedback.insert(INSERT, '⯈ Opening output directory')
#            feedback.insert(END, '\n\n') 
            startfile( parent_Dir + '\\' + folder_name_2 + '\\' + folder_name_3 )
            
#            print('_____________________________________________________\n')
#            print('3',os.getcwd())
#            print('_____________________________________________________\n\n')
            
        elif radio_outformat == 'Simple Workbook':
            '''
            Combined & Simple XLSX file
            '''
            
            ## Fn to combine
            Combine_XLSX(file_type,
                                             DB_index_col,
                                             file_n,
                                             parent_Dir,
                                             folder_name_2,
                                             folder_name_3,
                                             folder_name_7)
            ## Empty sheets dir
            Clear_Directory(parent_Dir, folder_name_2, folder_name_3)
            
            ## Open directory in Windows
#            feedback.insert(INSERT, '⯈ Opening output directory')
#            feedback.insert(END, '\n\n') 
            startfile( parent_Dir + '\\' + folder_name_2 + '\\' + folder_name_7 )
            
#            print('_____________________________________________________\n')
#            print('3',os.getcwd())
#            print('_____________________________________________________\n\n')
            
            
             
        else:
            pass
        
            
        try:
            print('======================================================================')
            print(dname + '\\' + folder_name_1)
            chdir(dname + '\\' + folder_name_1)
            print('======================================================================')
        except:
            pass
                         

    ##  (TEMP) All Results 
    print('---------------------------------------'                         )
    print('ALL Duplicate rows:\t',      all_duplicate_rows                  )
    print('ALL Erroneous ID\'s:\t',     all_no_err_ID                       )
    print('ALL Unverified ID\'s:\t',    all_unverify_ID                     )
    print('ALL Age blunders:\t',        all_mis_age                         )
    print('ALL Gendr blunders:\t',      all_mis_gender                      )
    print('ALL Citiz blunders:\t',      all_mis_citizen                     )
    print('ALL Youth blunders:\t',      all_mis_youth                       )
    print(                                                                  )
    print('ALL First Name:\t\t',        all_mis_name                        )
    print('ALL Surname:\t\t',           all_mis_sname                       )
    print('ALL Persal No.:\t\t',        all_mis_persal                      )
    print('ALL Salary Level:\t',        all_mis_SLvL                        )
    print('ALL Race:\t\t',              all_mis_race                        )
    print('ALL Disability:\t\t',        all_mis_disab                       )
    print('ALL Occ. Level:\t\t',        all_mis_occlvl                      )
    print('ALL Job Title:\t\t',         all_mis_jobtit                      )
    print('ALL Residence:\t\t',         all_mis_res                         )
    print(                                                                  )
    print('ALL Unverified T.I.:\t',     all_unveri_TI                       )
    print('ALL Interventions:\t',       all_intervention                    )
    print('ALL Type:\t\t',              all_TItype                          )
    print('ALL NQF:\t\t',               all_NQF                             )
    print('ALL duration:\t\t',          all_duration                        )
    print('ALL proivder:\t\t',          all_provider                        )
    print(                                                                  )
    print('ALL Res. Issues:\t',         all_res_issues                      )
    print(                                                                  )
    print('All Sheets:\t\t',            counter_all_sheets                  )
    print('All Books:\t\t',             counter_book                        )
    print('---------------------------------------'                         )      
    
    
    
    
    string_khangela_1 = ''
    string_khangela_2 = ''

    try:
        for jonga3 in set(all_khangela):
            string_khangela_1 = string_khangela_1 + '     •  ' + jonga3 + '\n'
        for jonga4 in khangela:
            string_khangela_2 = string_khangela_2 + '     •  ' + jonga4 + '\n'
    except UnboundLocalError:
        work_done = False
        print('\n\n>>> No files processed')
    
    

    if work_done == True:
        Complete_end    = time()
        DB_elapse_2     = str(round(Complete_end - DB_start))
        DB_elapse_s     = ('> Took {} seconds to complete all operations'.format(DB_elapse_2)) 
        
        ## Feedback Total elapsed time 
        feedback.insert(INSERT, '⯈ Took {} seconds to complete all operations'.format(DB_elapse_2))
        feedback.insert(END, '\n\n') 
        
        ## Feedback notification #6
        feedback.insert(INSERT, '⯈ Completed cleansing. Thank you for your patience :)')
    #                   + '\n\nPlease see Reports for remaining issues in output files'
    #                   + '\nTo see reports: Other → Result logs')
        feedback.insert(END, '\n\n') 
        
        
        
            
        if all_unverify_ID != 0 or all_no_err_ID != 0:
            if counter_book == 1:
                feedback.insert(INSERT, '⯈ Please review the following worksheet(s) containing\n     residual issues:')
                feedback.insert(END, '\n\n') 
                if counter_all_sheets == 1:
                    feedback.insert(INSERT, '{}'.format(string_khangela_2))
                    feedback.insert(END, '\n') 
                else:
                    feedback.insert(INSERT, '{}'.format(string_khangela_2))
                    feedback.insert(END, '\n') 
            else:
                feedback.insert(INSERT, '⯈ The following workbooks contain residual issues:\n\n{}'.format(string_khangela_1))
                feedback.insert(END, '\n') 
        else:
            pass
        
    
    else:
        feedback.insert(INSERT, '⯈ There were no workbooks to cleanse...  ¯\_(ツ)_/¯')
        feedback.insert(END, '\n\n') 
        
        
    feedback.yview_pickplace("end")


    

    

    '''
    Matplotlib 
    '''
    
    
    font = {'size' : 14}
    
    error_sources       = ['Unverified',
                           'Duplicates',
                           'ID', 
                           'Age', 
                           'Gender',
                           'Citizenship', 
                           'Youth', 
                           'First Name',
                           'Surname', 
                           'Persal', 
                           'Salary Level',
                           'Race', 
                           'Disabled',
                           'Occupation Lvl',
                           'Job Title',
                           'Residence',
                           'Unverified T.I.',
                           'T.I. Name',
                           'T.I. Type',
                           'T.I. NQF level',
                           'T.I. Duration',
                           'T.I. Provider']#,
#                           'Residual issues'
#                           ]
    
    

    
    
    
    error_source_values = [all_unverify_ID, 
                           all_duplicate_rows, 
                           all_no_err_ID, 
                           all_mis_age, 
                           all_mis_gender, 
                           all_mis_citizen, 
                           all_mis_youth, 
                           all_mis_name, 
                           all_mis_sname, 
                           all_mis_persal, 
                           all_mis_SLvL, 
                           all_mis_race, 
                           all_mis_disab,
                           all_mis_occlvl,
                           all_mis_jobtit,
                           all_mis_res,
                           all_unveri_TI,
                           all_intervention,
                           all_TItype,
                           all_NQF,
                           all_duration,
                           all_provider]#,
#                           all_res_issues
#                           ]
    
    error_array = array(error_source_values)
   
    
    
    
    '''
        (1) Bar Graph
    '''
    
    string_files = ''
    s_c = 0
    for sf in file_list:
        s_c += 1
        if sf == file_list[:-1]:
            string_files += sf
        elif s_c % 3 == 0:
            string_files = string_files + sf + ',\n'
        else:
            string_files = string_files + sf + ', '
            
            

#    try:
    " Console"
    #period_names = ['6 Months','1 Year','3 Years','5 Years']

#    bar_colours = ['#000000', 
#                   '#000000', 
#                   '#000000', 
#                   '#000000', 
#                   '#000000', 
#                   '#000000', 
#                   '#000000', 
#                   '#000000', 
#                   '#000000', 
#                   '#000000', 
#                   '#000000', 
#                   '#000000', 
#                   '#000000',
#                   '#000000',
#                   '#000000',
#                   '#000000',
#                   '#000000',
#                   '#000000',
#                   '#000000',
#                   '#000000',
#                   '#000000',
#                   '#000000'
#                    ]

    bar_colours = ['#ff9999',
                   '#ffaf99',
                   '#ffc799',
                   '#ffdb99',
                   '#fff899',
                   '#e7ff99',
                   '#c9ff99',
                   '#adff99',
                   '#99ffa7',
                   '#99ffce',
                   '#99ffec',
                   '#99eeff',
                   '#99d3ff',
                   '#99afff',
                   '#9e99ff',
                   '#b899ff',
                   '#ce99ff',
                   '#e799ff',
                   '#ff99fc',
                   '#ff99e6',
                   '#ff99c9',
                   '#ff99a8']#,
#                   '#ff9999'
#                   ]
                    
                    
    
#    bar_colours = cm.hsv(error_array/float(max(error_array)))
    
    plt.bar(error_sources, error_source_values, color= bar_colours, edgecolor ='black')
    plt.xticks(fontsize=8, rotation = 45)
    plt.yticks(fontsize=8, rotation = 0)
    # Matplotlib:
    plt.rcdefaults()
    #plt.style.use('ggplot')
    
    plt.grid(True)
    
    plt.xlabel('Sources of errors')
    plt.ylabel('Occurances')
    

    plt.gca().set_position((.1, .3, .8, .6))
    plt.figtext(.02, .02, "Workbooks represented here:\n\n" + string_files, color = 'black' , size = 6)
#    
    if counter_book > 1:
        tot_wrkbks = ' workbooks)'
    else:
        tot_wrkbks = ' workbook)'
                
    
    
    plt.title('All frequencies of errors\n(for ' + str(counter_book) + tot_wrkbks, fontweight='bold')
    #plt.savefig('{}.jpg'.format(fund_name))
    
    plt.grid(b=None)
    for i, v in enumerate(error_source_values):
        plt.text(error_sources[i], v+3 , str(v), color=bar_colours[i], fontweight='bold')
    
   
    

    #plt.style.use('dark_background')
    
    figdest = (parent_Dir +'\\'''+ folder_name_6 + '\\''Errors Bar Graph '+datetime.now().strftime('%d-%m-%Y at %H.%M')+'.png')
    plt.savefig(figdest)

    
#    plt.show()
    plt.close()
    
#    except:
#        pass

    
    
    
    "Tkinter"
    if radio_graphic2 == 'Active':
        try:
            Slave = Toplevel()
            Slave.geometry("1200x500")
            Slave.iconbitmap(image_dir+'Icon0.ico')
            #Slave.attributes('-fullscreen', True)

            
            Slave.geometry("%dx%d+0+0" % (win_w, win_h))
            
            f = Figure(figsize=(10,10), dpi=100)
            ax = f.add_subplot(111)
            ax.set_xlabel('Sources of errors', fontweight='bold')
            ax.set_ylabel('Occurances', fontweight='bold')
            ax.set_title('All frequencies of errors\n(for ' + str(counter_book) + tot_wrkbks, fontweight='bold')
            
            ax.set_facecolor('white')
            ax.set_navigate(True)
            ax.set_navigate_mode(True)
        
            
            rects1 = ax.bar(error_sources, error_source_values, 0.5, facecolor='#02319A')
            ax.tick_params(which='major', labelsize=5)
            
            
            canvas = FigureCanvasTkAgg(f, master=Slave)
            canvas.draw()
            canvas.get_tk_widget().pack(side=TOP, fill=BOTH, expand=1)
            
            toolbar = NavigationToolbar2Tk(canvas, Slave)
            toolbar.update()
            
            canvas.get_tk_widget().pack(side=TOP, fill=BOTH, expand=1)
        except:
            pass
    else:
        pass
    
    

#    '''
#        (2) Pie chart 1
#        
#        
#            Research:
#            SORT BY DESCENDING 
#            SAME FOR BAR
#    '''
#    fig, ax = plt.subplots(figsize=(6, 3), subplot_kw=dict(aspect="equal"))
#
#    
#    def func(pct, allvals):
#        absolute = int(pct/100.*np.sum(allvals))
#        return "{:.1f}%\n({:d})".format(pct, absolute)
#    
#    
#    wedges, texts, autotexts = ax.pie(error_source_values, autopct=lambda pct: func(pct, error_source_values),
#                                      textprops=dict(color="w"), counterclock = False, startangle = 0, shadow = True)
#    
#    ax.legend(wedges, error_sources,
#              title="Sources of errors",
#              loc="center left",
#              bbox_to_anchor=(1, 0, 0.5, 1))
#    
#    plt.setp(autotexts, size=4)#, weight="bold")
#    
#    ax.set_title("Occurance of errors in worksheet(s)")
#    
#    plt.show()


    '''
        (2) Pie chart 2

    '''





    '''
    Dirt Report
    '''

    Dirt_Report(parent_Dir, 
                folder_name_8,
                all_duplicate_rows,
                all_no_err_ID,
                all_unverify_ID,
                all_mis_age,
                all_mis_gender,
                all_mis_citizen,
                all_mis_youth,
                all_mis_name,
                all_mis_sname,
                all_mis_persal,
                all_mis_SLvL,
                all_mis_race,
                all_mis_disab,
                counter_all_sheets,
                counter_book,
                string_files,
                'All books',
                1,
                all_khangela,
                all_mis_occlvl,
                all_mis_jobtit,
                all_mis_res,
                all_unveri_TI,
                all_intervention,
                all_TItype,
                all_NQF,
                all_duration,
                all_provider,
                all_res_issues
                )
    
    

        
    Cleansing_Complete(counter_book)
    
    
    ## For some weird reason the working directory chanes to "Input files" after processessing 
    ## The working dir is reset here: 
    
    
#    print('_____________________________________________________\n')
#    print('4',os.getcwd())
#    print('_____________________________________________________\n\n')
    
    try:
        chdir(dname)
    except:
        pass
#    try:
#        os.chdir(os.getcwd().replace('Input files',''))
#    except:
#        feedback.insert(INSERT, '\n' + '[ERROR 0001] - Working directory failed to reset! Please RESTART application.')
#        feedback.insert(END, '\n\n')  
    
#    print('_____________________________________________________\n')
#    print('5',os.getcwd())
#    print('_____________________________________________________\n\n')
#    
    



# =============================================================================
# =============================================================================
# =============================================================================





def Verify_Write(df_db,
                 df_db_TI,
                 df, 
                 file_name, 
                 sheet_name, 
                 parent_Dir, 
                 folder_name_2,
                 folder_name_3,
                 folder_name_8,
                 file_type, 
                 duplicate_rows, 
                 Dup_Check_1, 
                 Dup_Check_2, 
                 Dup_Check_3,
                 cols_verify
                 ): 

    '''
    This function does all the Data Scrubbing 
    > Compares cell values to those in baseline data
    > Applies necessary corrections 
    > Saves number of errors detected 

    '''
    ## Columns to verify from ID number
    Age_col         = 'AGE'
    Gender_col      = 'GENDER'
    Youth_col       = 'YOUTH'
    Citizen_col     = 'NON-RSA CITIZEN'
    
    ## Columns to verify from Persal (Ref) DF
    ## fname, sname, persal, SLvL, IDno, OccLvL, res, race, disab
    ## [Persal DF, QMR, CQMR]



    col_fname           = ['FIRST NAMES',            Dup_Check_2,                           'NAMES OF THE LEARNER'      ]
    col_sname           = ['Surname',                Dup_Check_3,                           'SURNAME OF THE LEARNER'    ]
    col_persal          = ['Persalno',              'PERSAL NUMBER'                                                     ]
    col_SLvL            = ['SL-Person',             'SALARY LEVEL'                                                      ]
    col_IDno            = ['ID Number',             'ID NUMBER OF THE LEARNER',             'ID NUMBER OF THE LEARNER'  ]
    col_OccLvL          = ['Occ Cat Labour',        'OCCUPATIONAL LEVEL'                                                ]
    col_res             = ['ADDRESS 3',             'SPECIFY LEARNER RESIDENTIAL  AREA'                                 ]
    col_race            = ['Race',                  'RACE'                                                              ]
    col_disab           = ['Disabled',              'DISABILITY (YES/NO)',                  'DISABILITY'                ]
    col_job1            = ['Job Title on Rank',     'JOB TITLE'                                                         ]
    
    
    ## Note! For Training Interventions and Residual Columns - only catering for Latest format of QMR column header names
    col_interven        = ['Intervention',          'TRAINING INTERVENTION'                                             ]
    col_type            = ['Type',                  'TYPE OF LEARNING PROGRAMME'                                        ]
    col_NQF             = ['NQF Level',             'NQF LEVEL'                                                         ]
    col_duration        = ['Duration',              'DURATION OF PROGRAMME',                'Unit Of Measure'           ]
    col_provider        = ['Provider / Source',     'NAME OF THE TRAINING PROVIDER'                                     ]
    
    
    ## OFO codes column
    col_OFO             = ['OFO CODES (FOR OFFICE USE)'                                                                 ]
    
    ## Columns that do not have any reference data
    col_date_entered    = ['DATE ENTERED'                                                                               ]
    col_date_completed  = ['DATE COMPLETED'                                                                             ]
    col_amount_spent    = ['AMOUNT SPENT PER LEARNER (DIRECT COSTS-COURSE REGISTRATION)'                                ]
    col_venue           = ['Venue (Conference Facilities)'                                                              ]
    col_transport       = ['Transport (Indirect costs)'                                                                 ]
    col_accommodation   = ['Accommodation (Indirect costs)'                                                             ]
    col_catering        = ['Catering (Indirect costs)'                                                                  ]
    col_SETA_funded     = ['Is the Programme SETA/Industry Funded'                                                      ]
    col_completed       = ['COMPLETED/YES OR NO'                                                                        ]
    col_date_cert       = ['DATE THE LEARNER ISSUED WITH CERTIFICATE'                                                   ]
#    col_cert_number     = ['CERTIFICATE NUMBER'                                                                        ]                                           
    col_train_guide     = ['TRAINING GUIDELINE (WSP PRIORITIES)'                                                        ]
    col_train_facil     = ['NAME OF THE TRAINING FACILITATOR'                                                           ]
#    col_train_accred    = ['TRAINING PROVIDER ACREDITATION NUMBER'                                                     ]
    col_train_contact   = ['TRAININING PROVIDER CONTACT DETAILS'                                                        ]
    col_train_public    = ['IS TRAINING PROVIDER PRIVATE /PUBLIC (YES/NO)'                                              ]
    col_learner_local   = ['LEARNER LOCAL/DISTRICT MUNICIPALITY '                                                       ]
    col_learner_urban   = ['IS THE LEARNER RESIDENTIAL AREA URBAN / RURAL(YES/NO)'                                      ]
    col_municipality    = ['Muncipality (use Dropdown list worksheet)'                                                  ]
    

    
     
    
       
    
    ## Constructs appended to new file
    New_Prefix      = ''
    New_Suffix      = ''
    
    

    
    ## Local Mistake Tallies
    Glob_mis_age        = 0
    Glob_mis_gender     = 0
    Glob_mis_citizen    = 0
    Glob_mis_youth      = 0
    
    Glob_mis_name       = 0
    Glob_mis_sname      = 0
    Glob_mis_persal     = 0
    Glob_mis_SLvL       = 0
    Glob_mis_race       = 0
    Glob_mis_disab      = 0
    Glob_mis_occlvl     = 0
    Glob_mis_jobtit     = 0
    Glob_mis_res        = 0
    
    Glob_unveri_TI      = 0
    Glob_intervention   = 0
    Glob_TItype         = 0
    Glob_NQF            = 0
    Glob_duration       = 0
    Glob_provider       = 0
    
    Glob_res_issues     = 0 

    

    ## list of erronous ID numbers
    Blunder_ID      = []
    ## Unverified ID numbers (DNE in Persal)
    Unverified_ID   = []
    ## Unverified Training Interventions
    Unverified_TI   = []
    ## ID numbers /w Unverified Training Interventions
    unveri_TI_ID    = []
    ## Verified Training Interventions
    Verified_TI     = []
    ## QMR disabled value
    QMR_disab       = []
    
    
    ## lists of blunders detected within other columns
    blunder_col_date_entered        =   [] 
    blunder_col_date_completed      =   []
    blunder_col_amount_spent        =   []
    blunder_col_venue               =   []
    blunder_col_transport           =   []
    blunder_col_accommodation       =   []
    blunder_col_catering            =   []
    blunder_col_SETA_funded         =   []
    blunder_col_completed           =   []
    blunder_col_date_cert           =   []
#    blunder_col_cert_number         =   []                               
    blunder_col_train_guide         =   []
    blunder_col_train_facil         =   []
#    blunder_col_train_accred        =   []
    blunder_col_train_contact       =   []
    blunder_col_train_public        =   []
    blunder_col_learner_local       =   []
    blunder_col_learner_urban       =   []
    blunder_col_municipality        =   []
    
    
    '''
    Check for Duplicate entries in column
    '''
    ## list of duplicate Persal numbers
    ## QMR does not have persal, therefore: (try, except)
    try:
        Dup_Persal = []  
        for xx in df[Dup_Check_1]:
            chance = 0
            for yy in df[Dup_Check_1]:
                if xx == yy:
                    if yy in Dup_Persal:
                        pass
                    else:
                        # first duplicate detected (possibly- itself)
                        if chance == 0:
                            chance += 1
                        # second duplicate detected 
                        elif chance == 1:
                            Dup_Persal.append(yy)
                        # >1 duplicate(s) detected
                        else:
                            pass
                else:
                    pass
    except:
        pass
#    print(Dup_Persal)

#     ## Check on check for Duplicates 
#    check_dp = []
#    for aa in Dup_Persal:
#        c2 = 0
#        for bb in Dup_Persal:
#            if aa == bb:
#                if bb in check_dp:
#                    pass
#                else:
#                    if c2 == 0:
#                        c2 += 1
#                    elif c2 == 1:
#                        check_dp.append(bb)
#                    else:
#                        pass
#            else:
#                pass
#    print(check_dp) #list should be empty

    
    
    ## Iterate through each row of sheet 
    for index, row in df.iterrows():
        #ID_number   = str(round(int(index))) # INT ISSUE WHEN DOB >=2000
        ID_number   = str(index)
        

        ## Seperate ID number values
        try:
            DOB         = ID_number[0:6]
            DOB_year    = int(DOB[0:2])
            DOB_month   = int(DOB[2:4])
            DOB_day     = int(DOB[4:6])
            #print(index, ID_number, DOB, DOB_year, DOB_month, DOB_day)
            
            Gender      = int(ID_number[6:10])
            Citizen     = int(ID_number[10:11])
            Race        = int(ID_number[11:12])
            Checksum    = int(ID_number[12:13])

        except: 
            Blunder_ID.append(ID_number)
        
        '''
        Qualitative descriptions
        '''
        ## Skip this if ID is erroneous
        if ID_number in Blunder_ID:
            pass
        
        else:
            ## Full year of birth
            if DOB_year < 10:
                DOB_year += 2000
            else:
                DOB_year += 1900
            
            
            
            try:    
                Age = Calc_Age(date(DOB_year, DOB_month, DOB_day))
            except ValueError:
                ## Found in cases where:
                ##  ID number = 13digits, but month out of range (or)
                ##  ID number > 13 digits
                Blunder_ID.append(ID_number)
    
    
            ## Gender
            Q_Gender = ''
            if Gender >= 5000:
                Q_Gender = 'Male'
            else:
                Q_Gender = 'Female'
            
            ## Citizenship
            Q_Citizen = ''
            if Citizen == 0:
                Q_Citizen = 'RSA Citizen'
            else:
                Q_Citizen = 'NON-RSA Citizen'   
                
            ## Youth
            Q_Youth = ''
            if Age > 35:
                Q_Youth = 'No'
            else:
                Q_Youth = 'Yes'  
            
            '''
            Verify Sheet Data
            '''
            ## Verify from ID
            ## Verify Age
            if row[Age_col] != Age:
                df.at[index, Age_col] = Age
                Glob_mis_age += 1
            else:
                pass
            
            ## Verify Gender
            if str(row[Gender_col]).lower() != Q_Gender.lower():
                df.at[index, Gender_col] = Q_Gender
                Glob_mis_gender  += 1
            else:
                pass
            
            ## Verify Citizenship
            "fix 13/01/2020"
            "try/ except to bypass check if citizen column does not exist in test sheet "
            try:
                if str(row[Citizen_col]).lower() != Q_Citizen.lower():
                    df.at[index, Citizen_col] = Q_Citizen
                    Glob_mis_citizen  += 1
                else:
                    pass
            except:
                pass
            
            ## Verify Youth
            "fix 13/01/2020"
            try:
                if str(row[Youth_col]).lower() != Q_Youth.lower():
                    df.at[index, Youth_col] = Q_Youth
                    Glob_mis_youth  += 1
                else:
                    pass
            except:
                pass
            
              
            '''
            Verify data from Persal DF
            
                Redundant code ⯈ process repeated in an inelegant way
                Method, Verify_Persal(), available but needs finxing
            '''
            
            ## Verify from Persal DB                
            ##      Exception 1: QMR has different column names representing the same data 
            ##      Exception 2: Index D.N.E. in persal reference DF
            
            
            ## Verify F.Name
            ## Convert to method??? - 
            ## Use dictionary instead??? 
            
            #df = Verify_Persal(df_db, df, row, index, ID_number, cols_verify)
      

            
            ## F.Name 
            try:
                df_db_fname = sub(r"(\w)([A-Z])",r"\1 \2", (df_db.iloc[df_db.index.get_loc(int(ID_number))][col_fname[0]]).title().replace(' ',''))
                if row[col_fname[1]].title()   !=  df_db_fname:
                    #print(row[col_fname[1]], '\t\t' , df_db_fname)
                    df.at[index, col_fname[1]] = df_db_fname
                    Glob_mis_name += 1
                else:
                    pass
            except:
                try:
                    df_db_fname = sub(r"(\w)([A-Z])",r"\1 \2", (df_db.iloc[df_db.index.get_loc(int(ID_number))][col_fname[0]]).title().replace(' ',''))
                    if sub(r"(\w)([A-Z])",r"\1 \2", row[col_fname[2]].title().replace(' ',''))   !=  df_db_fname:
                        #print(row[col_fname[2]],'\t\t' ,  df_db_fname)
                        df.at[index, col_fname[2]] = df_db_fname
                        Glob_mis_name += 1
                    else:
                        pass
                except:
                    " ID is not found in Persal DF"
                    "   only need to append first case"
                    Unverified_ID.append(ID_number)
                    #print(index,'D.N.E in ref df')
                    
                    
                    

            
            
            
            ## S.Name 
            try:
                df_db_sname = sub(r"(\w)([A-Z])",r"\1 \2", (df_db.iloc[df_db.index.get_loc(int(ID_number))][col_sname[0]]).title().replace(' ',''))
                if row[col_sname[1]].title()   !=  df_db_sname:
                    #print(row[col_sname[1]], '\t\t' , df_db_sname)
                    df.at[index, col_sname[1]] = df_db_sname
                    Glob_mis_sname += 1
                else:
                    pass
            except:
                try:
                    df_db_sname = sub(r"(\w)([A-Z])",r"\1 \2", (df_db.iloc[df_db.index.get_loc(int(ID_number))][col_sname[0]]).title().replace(' ',''))
                    if sub(r"(\w)([A-Z])",r"\1 \2", row[col_sname[2]].title().replace(' ',''))   !=  df_db_sname:
                        #print(row[col_sname[2]],'\t\t' ,  df_db_sname)
                        df.at[index, col_sname[2]] = df_db_sname
                        Glob_mis_sname += 1
                    else:
                        pass
                except:
                    #print(index,'D.N.E.')
                    pass
            
            
            
            ## Persal 
            try:
                df_db_persal = df_db.iloc[df_db.index.get_loc(int(ID_number))][col_persal[0]]
                
                if str(row[col_persal[1]])  !=  str(df_db_persal):
                    #print('fail\t\t', row[col_persal[1]], '\t\t' , df_db_persal)
                    df.at[index, col_persal[1]] = df_db_persal 
                    Glob_mis_persal += 1
                else:
                    #print('pass')
                    pass
            except:
                pass
            


            ## SLvL
            try:
                df_db_SLvL = df_db.iloc[df_db.index.get_loc(int(ID_number))][col_SLvL[0]]
                #print(row[col_SLvL[1]], '\t\t' , df_db_SLvL)
                if row[col_SLvL[1]]   !=  df_db_SLvL:
                    #print(row[col_SLvL[1]], '\t\t' , df_db_SLvL)
                    df.at[index, col_SLvL[1]] = df_db_SLvL
                    Glob_mis_SLvL += 1
                else:
                    pass
            except:
                pass
                


            # OccLvL
            try:
               df_db_OccLvL = sub(r"(\w)([A-Z])",r"\1 \2", (df_db.iloc[df_db.index.get_loc(int(ID_number))][col_OccLvL[0]]).title().replace(' ',''))
               #print(row[col_OccLvL[1]], '\t\t' , df_db_OccLvL)
               if row[col_OccLvL[1]]   !=  df_db_OccLvL:
                   #print(row[col_OccLvL[1]], '\t\t' , df_db_OccLvL)
                   df.at[index, col_OccLvL[1]] = df_db_OccLvL
                   Glob_mis_occlvl += 1
               else:
                   pass
            except: 
               pass 





            ## Res 
            try:
                df_db_res = sub(r"(\w)([A-Z])",r"\1 \2", (df_db.iloc[df_db.index.get_loc(int(ID_number))][col_res[0]]).title().replace(' ',''))
                #print(row[col_res[1]], '\t\t' , df_db_res)
                if row[col_res[1]].title()   !=  df_db_res:
                    df.at[index, col_res[1]] = df_db_res
                    Glob_mis_res += 1
                else:
                    pass
            except:
                pass
            
            

            ## Race            
            try:
                df_db_race = sub(r"(\w)([A-Z])",r"\1 \2", (df_db.iloc[df_db.index.get_loc(int(ID_number))][col_race[0]]).title().replace(' ',''))
                #print(row[col_race[1]], '\t\t' , df_db_race)
                if row[col_race[1]].title()   !=  df_db_race:
                    #print(row[col_race[1]], '\t\t' , df_db_race)
                    df.at[index, col_race[1]] = df_db_race
                    Glob_mis_race += 1
                else:
                    pass
            except:
                pass
            
            
            
            ## Disab
            ' Disab replaces QMR '
#            try:
#                df_db_disab = sub(r"(\w)([A-Z])",r"\1 \2", (df_db.iloc[df_db.index.get_loc(int(ID_number))][col_disab[0]]).title().replace(' ',''))
#                
#                if df_db_disab.lower() == 'n':
#                    df_db_disab = 'No'
#                else:
#                    df_db_disab = 'Yes'  
#                
#                #print(row[col_disab[1]].title(), '\t\t' , df_db_disab)
#                
#                if row[col_disab[1]].title()   !=  df_db_disab:
#                    #print(row[col_disab[1]].title(), '\t\t' , df_db_disab)
#                    df.at[index, col_disab[1]] = df_db_disab
#                    Glob_mis_disab += 1
#                else:
#                    pass
#            except:
#                try:
#                    df_db_disab = sub(r"(\w)([A-Z])",r"\1 \2", (df_db.iloc[df_db.index.get_loc(int(ID_number))][col_disab[0]]).title().replace(' ',''))
#                    
#                    if df_db_disab.lower() == 'n':
#                        df_db_disab = 'No'
#                    else:
#                        df_db_disab = 'Yes'  
#                        
#                    #print(row[col_disab[2]].title(), '\t\t' , df_db_disab)
#                    
#                    if row[col_disab[2]].title()   !=  df_db_disab:
#                        #print(row[col_disab[2]].title(), '\t\t' , df_db_disab)
#                        df.at[index, col_disab[2]] = df_db_disab
#                        Glob_mis_disab += 1
#                        
#                
#                    else:
#                        pass
#                except:
#                    pass
            ## Disab
            ' Disab keeps QMR '
            '   Appends to list for special cell formatting later'
            try:
                df_db_disab = sub(r"(\w)([A-Z])",r"\1 \2", (df_db.iloc[df_db.index.get_loc(int(ID_number))][col_disab[0]]).title().replace(' ',''))
                
                if df_db_disab.lower() == 'n':
                    df_db_disab = 'No'
                else:
                    df_db_disab = 'Yes'  
                
                if row[col_disab[1]].title()   !=  df_db_disab:
                    QMR_disab.append(ID_number)
                else:
                    pass
            except:
                try:
                    df_db_disab = sub(r"(\w)([A-Z])",r"\1 \2", (df_db.iloc[df_db.index.get_loc(int(ID_number))][col_disab[0]]).title().replace(' ',''))
                    
                    if df_db_disab.lower() == 'n':
                        df_db_disab = 'No'
                    else:
                        df_db_disab = 'Yes'  
                        
                    if row[col_disab[2]].title()   !=  df_db_disab:
                        QMR_disab.append(ID_number)
                    else:
                        pass
                except:
                    pass
            
                        
            
            ## Job Title (in QMR)
            try:
                df_db_job1 = sub(r"(\w)([A-Z])",r"\1 \2", (df_db.iloc[df_db.index.get_loc(int(ID_number))][col_job1[0]]).title().replace(' ',''))
                if row[col_job1[1]].title()   !=  df_db_job1:
                    #print(row[col_job1[1]], '\t\t' , df_db_job1)
                    df.at[index, col_job1[1]] = df_db_job1
                    Glob_mis_jobtit += 1
                else:
                    pass
#            except Exception as ex:
#                print(row[col_job1[1]].title() )
#                template = "An exception of type {0} occurred. Arguments:\n{1!r}"
#                message = template.format(type(ex).__name__, ex.args)
#                print(message)
            except:
                # DNE
                pass








            '''
            Other Checks - Columns that cannot be checked against Reference Data
                
            
                Note:   Does not detect any variations in column header names, 
                        only one header name is used [0]
                        
                        if script is unable to find the column header name:
                        it handles as a KeyError exception in the method 
            
            '''
                                    
            ## Amount columns
            'Other_Cols_Check_1'
            blunder_col_amount_spent,   Glob_res_issues     = Other_Cols_Check_1(Glob_res_issues, row, col_amount_spent[0],  index, blunder_col_amount_spent )
            blunder_col_venue,          Glob_res_issues     = Other_Cols_Check_1(Glob_res_issues, row, col_venue[0],         index, blunder_col_venue        )
            blunder_col_transport,      Glob_res_issues     = Other_Cols_Check_1(Glob_res_issues, row, col_transport[0],     index, blunder_col_transport    )
            blunder_col_accommodation,  Glob_res_issues     = Other_Cols_Check_1(Glob_res_issues, row, col_accommodation[0], index, blunder_col_accommodation)
            blunder_col_catering,       Glob_res_issues     = Other_Cols_Check_1(Glob_res_issues, row, col_catering[0],      index, blunder_col_catering     )

                    
            ## Binary columns
            'Other_Cols_Check_2'
            blunder_col_SETA_funded,    Glob_res_issues     = Other_Cols_Check_2(Glob_res_issues, row, index, col_SETA_funded[0],    blunder_col_SETA_funded,    'seta',     'industry'              )
            blunder_col_completed,      Glob_res_issues     = Other_Cols_Check_2(Glob_res_issues, row, index, col_completed[0],      blunder_col_completed,      'yes',      'no'                    )
            blunder_col_train_public,   Glob_res_issues     = Other_Cols_Check_2(Glob_res_issues, row, index, col_train_public[0],   blunder_col_train_public,   'public',   'private'               )
            blunder_col_learner_local,  Glob_res_issues     = Other_Cols_Check_2(Glob_res_issues, row, index, col_learner_local[0],  blunder_col_learner_local,  'local',    'district municipality' )
            blunder_col_learner_urban,  Glob_res_issues     = Other_Cols_Check_2(Glob_res_issues, row, index, col_learner_urban[0],  blunder_col_learner_urban,  'urban',    'rural'                 )
            
            
            ## Date certificate issued 
            'Other_Cols_Check_3'
            blunder_col_municipality,   Glob_res_issues    = Other_Cols_Check_3(Glob_res_issues, df, index, row,
                                                              col_date_cert[0], blunder_col_date_cert,
                                                              'not', 'still')
            ## Date entered 
            blunder_col_date_entered,   Glob_res_issues    = Other_Cols_Check_3(Glob_res_issues, df, index, row,
                                                              col_date_entered[0], blunder_col_date_entered,
                                                              'Not_needed', 'Not_needed')
            ## Date completed
            blunder_col_date_completed, Glob_res_issues  = Other_Cols_Check_3(Glob_res_issues, df, index, row,
                                                              col_date_completed[0], blunder_col_date_completed,
                                                              'Not_needed', 'Not_needed')
            
            ## Training contact details
            ##  email
            try:
                if  '@' in str(row[col_train_contact[0]]):
                    pass
                ##  phone number
                elif isinstance(float(str(row[col_train_contact[0]]).lower().replace(' ','').replace(',','').replace("(",'').replace(')','').replace('+','').replace("/",'').replace('-','')), float):
                    pass
                else:
                    blunder_col_train_contact.append(index)
                    Glob_res_issues += 1
            except ValueError:
                blunder_col_train_contact.append(index)
                Glob_res_issues += 1
            except KeyError:
                ## Assume that col header name not accounter for, or DNE
                print('KeyError (Check for training contact details)')

                              
            ## Checks Training Guideline, Training Facilitator, Municipality 
            'Weak - only checks if cells are not blank'
            
            try:
                if isnull(row[col_train_guide[0]]) is True:
                    blunder_col_train_guide.append(index)
                    Glob_res_issues += 1
                else:
                    pass
                
                if isnull(row[col_train_facil[0]]) is True:
                    blunder_col_train_facil.append(index)
                    Glob_res_issues += 1
                else:
                    pass
                
                if isnull(row[col_municipality[0]]) is True:
                    blunder_col_municipality.append(index)
                    Glob_res_issues += 1
                else:
                    pass
            except KeyError:
                ## Assume that col header name not accounter for, or DNE
                print('KeyError (Checks for training guidline, facilitator and muni.)')
            
            
           
  
            
            
            '''
            Verify data from Training Interteventions DF

            '''
            try:

                TI_Index = row[col_interven[1]]
                                
                Glob_TItype, Glob_NQF, Glob_provider, Glob_duration, Verified_TI, Unverified_TI, unveri_TI_ID = Training_Intervention_Verifier(index, row, TI_Index,        
                                                                                                                                 
                                                                                                                                     df,              df_db_TI,
                                                                                                                               
                                                                                                                                   col_type,        Glob_TItype,
                                                                                                                                   col_NQF,         Glob_NQF,
                                                                                                                                   col_provider,    Glob_provider,
                                                                                                                                   col_duration,    Glob_duration,
                                                                                                                                   
                                                                                                                                   Verified_TI,     Unverified_TI,
                                                                                                                                                    unveri_TI_ID
                                                                                                                               )
            except KeyError:
                ## Assumption that 'TRAINING INTERVENTION' DNE
                pass
 
    Unverified_TI   = set(Unverified_TI)
    Glob_unveri_TI  = len(Unverified_TI)
    Verified_TI     = set(Verified_TI)
    len_Veri_TI     = len(Verified_TI)
  
    'Temporary text reports on TI verified'
    'Stored in input files dir'
#    Unveri_txt  = open('Unverified Training Interventions - {}.txt'.format(sheet_name),'w')
#    Veri_txt    = open('Verified Training Interventions - {}.txt'.format(sheet_name),'w')
#    
#    Unveri_txt.write('{} Unverified Training Interventions:\n\n\n'.format(Glob_unveri_TI))
#    Veri_txt.write('{} Verified Training Interventions:\n\n\n'.format(len_Veri_TI))
#    
#    for unveri in Unverified_TI:
#        Unveri_txt.write(unveri+'\n')    
#    for veri in Verified_TI:
#        Veri_txt.write(veri+'\n')
#
#    Unveri_txt.close()
#    Veri_txt.close()
    
    
    bad_ID = Blunder_ID.copy()
    
        
    
    
    
    
    '''
    Check to see if erroneous ID numbers has duplicate row data   
    
    '''
        
    for blunder2 in Blunder_ID:
        ## Get Persal no. attached to erroneous ID 
        B_IDpersal = ''           
        try:
                        ## df.at[row, column]
            B_IDpersal  = (df.at[int(blunder2), Dup_Check_1])                       # FLOAT?
        except:
            pass
            
        ## Check if Persal no. is a duplicate
        if B_IDpersal in Dup_Persal:
            ## Get list of indices (ID no's) that share persal no.
            list_dup = df.index[df[Dup_Check_1]==B_IDpersal].tolist()
            
            ## Remove ID under observation (Blunder ID)
            list_dup.remove(float(blunder2))
            
            ## Check if other ID's share other common column values with Blunder ID
            name_check_1        = df.at[int(blunder2), Dup_Check_2]                 # FLOAT?
            sname_check_1       = df.at[int(blunder2), Dup_Check_3]                 # FLOAT?
            
            for vv in list_dup:
                ## Ensure ID compared to has valid length
                if len(str(int(vv))) == 13:
                    name_check_2    = df.at[vv, Dup_Check_2]
                    sname_check_2   = df.at[vv, Dup_Check_3]
                    
                    if name_check_1 + sname_check_1 == name_check_2 + sname_check_2:
                        ## drop row containing erroneous ID
                        df = df.drop(float(blunder2))
                        ## drop erroneous ID from list of erroneous ID's
                        Blunder_ID.remove(blunder2)
                        break 
                    else:
                        pass
                else:
                    pass                   
        else:
            pass
    
    
    
    
    '''
    fix for missing or incorrect IDs
    
        will only work if:
            S.Name, F.Name, Race in test sheet is correct 
            (I.e. test sheet matches persal sheet)
    '''
    
    for bID in bad_ID:
#        print('XXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXX')
#        print(bID)
        
        
        ## Format of F.Name & S.Name in Persal DF:
        ##      F.Name & S.Name take up 40 characters, including spaces
        sil_fname = 40
        sil_sname = 20
        
        try:
            ## Get name, surname, race from df
            val_test1   = df.at[bID, col_fname[1]]
            val_test2   = df.at[bID, col_sname[1]]
            val_test3   = df.at[bID, col_race[1]]
            
            ## Get format into that of persal df
            ## add spaces
            while len(val_test1) < sil_fname:
                val_test1 += ' '
            val_test1 = val_test1.upper()
            
            while len(val_test2) < sil_sname:
                val_test2 += ' '
            val_test2 = val_test2.upper()
            
            ## if all 3 cells match then an ID is found
            compare = df_db.loc[(df_db[col_fname[0]] == val_test1) & (df_db[col_sname[0]] == val_test2) & (df_db[col_race[0]] == val_test3)]
            compare = compare.index.tolist()
            
            ## replace the erronous/place-holder ID with the real ID number
            df = df.rename(index={ bID : str(compare[0]) })
            

            Blunder_ID.remove(bID)
            
        
        except:
            ## Event in which bID needs to be converted to integer
            try:
#                print('pass')
                val_test1   = df.at[int(bID), col_fname[1]]
                val_test2   = df.at[int(bID), col_sname[1]]
                val_test3   = df.at[int(bID), col_race[1]]
                
                while len(val_test1) < sil_fname:
                    val_test1 += ' '
                val_test1 = val_test1.upper()
                
                while len(val_test2) < sil_sname:
                    val_test2 += ' '
                val_test2 = val_test2.upper()
                
                compare = df_db.loc[(df_db[col_fname[0]] == val_test1) & (df_db[col_sname[0]] == val_test2) & (df_db[col_race[0]] == val_test3)]
                compare = compare.index.tolist()
#                print(compare)
                
                df = df.rename(index={ bID : str(compare[0]) })
                
    
                Blunder_ID.remove(bID)
            except:
                pass
        
#        print('XXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXX')
#        print()
    
    
    
    
    ## Write to new Excel file
    "File name includes parent workbook name"
#    file_Dir_2 = (parent_Dir 
#                  + '\\''' 
#                  + folder_name_2
#                  + '\\''' 
#                  + folder_name_3 
#                  + '\\''' 
#                  + New_Prefix 
#                  + file_name.replace(file_type.replace('*',''),'')
#                  + New_Suffix
#                  + ' ('
#                  + sheet_name
#                  + ') '
#                  + file_type.replace('*','')
#                  )    
    "File name only contains sheets name"
    ## Possible problem when multiple workbooks /w same sheetnames 
    print(file_type)
    file_Dir_2 = (parent_Dir 
                  + '\\''' 
                  + folder_name_2
                  + '\\''' 
                  + folder_name_3 
                  + '\\''' 
                  + sheet_name
                  + file_type.replace('*','')
                  )    
    
    writer = ExcelWriter(file_Dir_2, engine='xlsxwriter')
    df.to_excel(writer, sheet_name = sheet_name)
    workbook  = writer.book
    worksheet = writer.sheets[sheet_name]
       
    ## Add a header format.
    header_format = workbook.add_format(
            {
            'bold': True,
            'text_wrap': True,
            'align': 'bottom',
            'fg_color': '#1F487C',
            'font_color': 'white',
            'font_size': 9,
            'font_name': 'Century Gothic',
            'border': 1
            }
            )

    INDEX_format = workbook.add_format(
            {
            'bold': True,
            'num_format': '#',
            'align': 'centre'
            }
            )

    error_format = workbook.add_format(
            {
            'bold': False,
            'text_wrap': True,
            'fg_color': '#ff5c5c',
            'font_color': 'black',
            'font_size': 9,
            'font_name': 'Century Gothic',
            'border': 1,
            'align': 'left'
            }
            )
    
    #f0bf65'
    unverified_format = workbook.add_format(
            {
            'bold': False,
            'text_wrap': True,
            'fg_color': '#ff9563',
            'font_color': 'black',
            'font_size': 9,
            'font_name': 'Century Gothic',
            'border': 1,
            'align': 'left'
            }
            )
    
    
    unverified_TI_format = workbook.add_format(
            {
            'bold': False,
            'text_wrap': True,
            'fg_color': '#fff45c',
            'font_color': 'black',
            'font_size': 9,
            'font_name': 'Century Gothic',
            'border': 1,
            'align': 'left'
            }
            )
    
    disab_format = workbook.add_format(
        {
        'bold': False,
        'text_wrap': True,
        'fg_color': '#ff8ad8',
        'font_color': 'black',
        'font_size': 9,
        'font_name': 'Century Gothic',
        'border': 1,
        'align': 'left'
        }
        )
    
    
    
    ## Format - Global 
    workbook.formats[0].set_font_size(9)
    workbook.formats[0].set_font_name('Century Gothic')
    workbook.formats[0].set_border(1)
    #workbook.formats[0].set_border_color('#FF0000')
    

    ## Format col headers                              
    for col_num, value in enumerate(df.columns.values):
        worksheet.write(0, col_num +1 , value, header_format)   
    
    ## Format col width
    for col in range(len(df.columns) + 1): 
        worksheet.set_column(col, col, 20)
   
    ## Format Index
    for index, row in df.iterrows():
        #index = int(index)
        try:
            worksheet.write(df.index.get_loc(index)+1, 0, index, INDEX_format)
        except:
            pass


    missing_data    = 'Missing Data'
    missing_data_2  = '-'

    ## Highlight QMR Disabled value that differs from Persal
    for QMRDisab in QMR_disab:
        Format_Excel_Cell(df, index, worksheet, QMRDisab, col_disab[1],      1, missing_data_2, disab_format)




    
    
    ## Format unverified ID #s (No Persal records)
    for no_veri in Unverified_ID:
        try:
            worksheet.write(df.index.get_loc(float(no_veri)) +1 , 
                            0, 
                            str(round(float(no_veri))), unverified_format
                            ) #+1 account for header
            
            Format_Excel_Cell(df, index, worksheet, no_veri, col_fname[1],      1, missing_data_2, unverified_format)
            Format_Excel_Cell(df, index, worksheet, no_veri, col_sname[1],      1, missing_data_2, unverified_format)     
            Format_Excel_Cell(df, index, worksheet, no_veri, col_persal[1],     1, missing_data_2, unverified_format)     
            Format_Excel_Cell(df, index, worksheet, no_veri, col_SLvL[1],       1, missing_data_2, unverified_format)     
            Format_Excel_Cell(df, index, worksheet, no_veri, col_OccLvL[1],     1, missing_data_2, unverified_format)     
            Format_Excel_Cell(df, index, worksheet, no_veri, col_res[1],        1, missing_data_2, unverified_format) 
            Format_Excel_Cell(df, index, worksheet, no_veri, col_race[1],       1, missing_data_2, unverified_format)
            Format_Excel_Cell(df, index, worksheet, no_veri, col_disab[1],      1, missing_data_2, unverified_format)
            Format_Excel_Cell(df, index, worksheet, no_veri, col_job1[1],       1, missing_data_2, unverified_format)
            
            Format_Excel_Cell(df, index, worksheet, no_veri, Age_col,           1, missing_data_2, unverified_format) 
            Format_Excel_Cell(df, index, worksheet, no_veri, Gender_col,        1, missing_data_2, unverified_format)
            Format_Excel_Cell(df, index, worksheet, no_veri, Youth_col,         1, missing_data_2, unverified_format)
            Format_Excel_Cell(df, index, worksheet, no_veri, Citizen_col,       1, missing_data_2, unverified_format)
            

    
        except ValueError:
            worksheet.write(df.index.get_loc(no_veri) +1 , 
                            0, 
                            no_veri, unverified_format)
            
            
            Format_Excel_Cell(df, index, worksheet, no_veri, col_fname[1],      1, missing_data_2, unverified_format)
            Format_Excel_Cell(df, index, worksheet, no_veri, col_sname[1],      1, missing_data_2, unverified_format)     
            Format_Excel_Cell(df, index, worksheet, no_veri, col_persal[1],     1, missing_data_2, unverified_format)     
            Format_Excel_Cell(df, index, worksheet, no_veri, col_SLvL[1],       1, missing_data_2, unverified_format)     
            Format_Excel_Cell(df, index, worksheet, no_veri, col_OccLvL[1],     1, missing_data_2, unverified_format)     
            Format_Excel_Cell(df, index, worksheet, no_veri, col_res[1],        1, missing_data_2, unverified_format) 
            Format_Excel_Cell(df, index, worksheet, no_veri, col_race[1],       1, missing_data_2, unverified_format)
            Format_Excel_Cell(df, index, worksheet, no_veri, col_disab[1],      1, missing_data_2, unverified_format)
            Format_Excel_Cell(df, index, worksheet, no_veri, col_job1[1],       1, missing_data_2, unverified_format)
            
            Format_Excel_Cell(df, index, worksheet, no_veri, Age_col,           1, missing_data_2, unverified_format) 
            Format_Excel_Cell(df, index, worksheet, no_veri, Gender_col,        1, missing_data_2, unverified_format)
            Format_Excel_Cell(df, index, worksheet, no_veri, Youth_col,         1, missing_data_2, unverified_format)
            Format_Excel_Cell(df, index, worksheet, no_veri, Citizen_col,       1, missing_data_2, unverified_format)
            
        except KeyError:
            try:
                worksheet.write(df.index.get_loc(int(no_veri)) +1 , 
                                0, 
                                no_veri, unverified_format)   
                
                
                Format_Excel_Cell(df, index, worksheet, no_veri, col_fname[1],      1, missing_data_2, unverified_format)
                Format_Excel_Cell(df, index, worksheet, no_veri, col_sname[1],      1, missing_data_2, unverified_format)     
                Format_Excel_Cell(df, index, worksheet, no_veri, col_persal[1],     1, missing_data_2, unverified_format)     
                Format_Excel_Cell(df, index, worksheet, no_veri, col_SLvL[1],       1, missing_data_2, unverified_format)     
                Format_Excel_Cell(df, index, worksheet, no_veri, col_OccLvL[1],     1, missing_data_2, unverified_format)     
                Format_Excel_Cell(df, index, worksheet, no_veri, col_res[1],        1, missing_data_2, unverified_format) 
                Format_Excel_Cell(df, index, worksheet, no_veri, col_race[1],       1, missing_data_2, unverified_format)
                Format_Excel_Cell(df, index, worksheet, no_veri, col_disab[1],      1, missing_data_2, unverified_format)
                Format_Excel_Cell(df, index, worksheet, no_veri, col_job1[1],       1, missing_data_2, unverified_format)
                
                Format_Excel_Cell(df, index, worksheet, no_veri, Age_col,           1, missing_data_2, unverified_format) 
                Format_Excel_Cell(df, index, worksheet, no_veri, Gender_col,        1, missing_data_2, unverified_format)
                Format_Excel_Cell(df, index, worksheet, no_veri, Youth_col,         1, missing_data_2, unverified_format)
                Format_Excel_Cell(df, index, worksheet, no_veri, Citizen_col,       1, missing_data_2, unverified_format)
                
            except:
                try:
                    worksheet.write(df.index.get_loc(no_veri) +1 , 
                                    0, 
                                    no_veri, unverified_format)  
                
                    Format_Excel_Cell(df, index, worksheet, no_veri, col_fname[1],      1, missing_data_2, unverified_format)
                    Format_Excel_Cell(df, index, worksheet, no_veri, col_sname[1],      1, missing_data_2, unverified_format)     
                    Format_Excel_Cell(df, index, worksheet, no_veri, col_persal[1],     1, missing_data_2, unverified_format)     
                    Format_Excel_Cell(df, index, worksheet, no_veri, col_SLvL[1],       1, missing_data_2, unverified_format)     
                    Format_Excel_Cell(df, index, worksheet, no_veri, col_OccLvL[1],     1, missing_data_2, unverified_format)     
                    Format_Excel_Cell(df, index, worksheet, no_veri, col_res[1],        1, missing_data_2, unverified_format) 
                    Format_Excel_Cell(df, index, worksheet, no_veri, col_race[1],       1, missing_data_2, unverified_format)
                    Format_Excel_Cell(df, index, worksheet, no_veri, col_disab[1],      1, missing_data_2, unverified_format)
                    Format_Excel_Cell(df, index, worksheet, no_veri, col_job1[1],       1, missing_data_2, unverified_format)
                    
                    Format_Excel_Cell(df, index, worksheet, no_veri, Age_col,           1, missing_data_2, unverified_format) 
                    Format_Excel_Cell(df, index, worksheet, no_veri, Gender_col,        1, missing_data_2, unverified_format)
                    Format_Excel_Cell(df, index, worksheet, no_veri, Youth_col,         1, missing_data_2, unverified_format)
                    Format_Excel_Cell(df, index, worksheet, no_veri, Citizen_col,       1, missing_data_2, unverified_format)
                
                except:
                    print('\n[Unknown ERROR] can\'t locate unverified Index!: ',no_veri, '(',type(no_veri),')\n')
            
        

    
    
    
    
    ## Format erroneous ID #s   
    for blunder1 in Blunder_ID:
        try:
            worksheet.write(df.index.get_loc(float(blunder1)) +1 , 
                            0, 
                            str(round(float(blunder1))), error_format
                            ) #+1 account for header
            
            Format_Excel_Cell(df, index, worksheet, blunder1, col_fname[1],     1, missing_data_2, error_format)
            Format_Excel_Cell(df, index, worksheet, blunder1, col_sname[1],     1, missing_data_2, error_format)     
            Format_Excel_Cell(df, index, worksheet, blunder1, col_persal[1],    1, missing_data_2, error_format)     
            Format_Excel_Cell(df, index, worksheet, blunder1, col_SLvL[1],      1, missing_data_2, error_format)     
            Format_Excel_Cell(df, index, worksheet, blunder1, col_OccLvL[1],    1, missing_data_2, error_format)     
            Format_Excel_Cell(df, index, worksheet, blunder1, col_res[1],       1, missing_data_2, error_format) 
            Format_Excel_Cell(df, index, worksheet, blunder1, col_race[1],      1, missing_data_2, error_format)
            Format_Excel_Cell(df, index, worksheet, blunder1, col_disab[1],     1, missing_data_2, error_format)
            Format_Excel_Cell(df, index, worksheet, blunder1, col_job1[1],      1, missing_data_2, error_format)
            
            Format_Excel_Cell(df, index, worksheet, blunder1, Age_col,          1, missing_data_2, error_format) 
            Format_Excel_Cell(df, index, worksheet, blunder1, Gender_col,       1, missing_data_2, error_format)
            Format_Excel_Cell(df, index, worksheet, blunder1, Youth_col,        1, missing_data_2, error_format)
            Format_Excel_Cell(df, index, worksheet, blunder1, Citizen_col,      1, missing_data_2, error_format)
            
            

        except ValueError:
            worksheet.write(df.index.get_loc(blunder1) +1 , 
                            0, 
                            blunder1, error_format)
            
            Format_Excel_Cell(df, index, worksheet, blunder1, col_fname[1],     1, missing_data_2, error_format)
            Format_Excel_Cell(df, index, worksheet, blunder1, col_sname[1],     1, missing_data_2, error_format)     
            Format_Excel_Cell(df, index, worksheet, blunder1, col_persal[1],    1, missing_data_2, error_format)     
            Format_Excel_Cell(df, index, worksheet, blunder1, col_SLvL[1],      1, missing_data_2, error_format)     
            Format_Excel_Cell(df, index, worksheet, blunder1, col_OccLvL[1],    1, missing_data_2, error_format)     
            Format_Excel_Cell(df, index, worksheet, blunder1, col_res[1],       1, missing_data_2, error_format) 
            Format_Excel_Cell(df, index, worksheet, blunder1, col_race[1],      1, missing_data_2, error_format)
            Format_Excel_Cell(df, index, worksheet, blunder1, col_disab[1],     1, missing_data_2, error_format)
            Format_Excel_Cell(df, index, worksheet, blunder1, col_job1[1],      1, missing_data_2, error_format)
            
            Format_Excel_Cell(df, index, worksheet, blunder1, Age_col,          1, missing_data_2, error_format) 
            Format_Excel_Cell(df, index, worksheet, blunder1, Gender_col,       1, missing_data_2, error_format)
            Format_Excel_Cell(df, index, worksheet, blunder1, Youth_col,        1, missing_data_2, error_format)
            Format_Excel_Cell(df, index, worksheet, blunder1, Citizen_col,      1, missing_data_2, error_format)
            
        except KeyError:
            worksheet.write(df.index.get_loc(blunder1) +1 , 
                            0, 
                            blunder1, error_format)
            
            Format_Excel_Cell(df, index, worksheet, blunder1, col_fname[1],     1, missing_data_2, error_format)
            Format_Excel_Cell(df, index, worksheet, blunder1, col_sname[1],     1, missing_data_2, error_format)     
            Format_Excel_Cell(df, index, worksheet, blunder1, col_persal[1],    1, missing_data_2, error_format)     
            Format_Excel_Cell(df, index, worksheet, blunder1, col_SLvL[1],      1, missing_data_2, error_format)     
            Format_Excel_Cell(df, index, worksheet, blunder1, col_OccLvL[1],    1, missing_data_2, error_format)     
            Format_Excel_Cell(df, index, worksheet, blunder1, col_res[1],       1, missing_data_2, error_format) 
            Format_Excel_Cell(df, index, worksheet, blunder1, col_race[1],      1, missing_data_2, error_format)
            Format_Excel_Cell(df, index, worksheet, blunder1, col_disab[1],     1, missing_data_2, error_format)
            Format_Excel_Cell(df, index, worksheet, blunder1, col_job1[1],      1, missing_data_2, error_format)
            
            Format_Excel_Cell(df, index, worksheet, blunder1, Age_col,          1, missing_data_2, error_format) 
            Format_Excel_Cell(df, index, worksheet, blunder1, Gender_col,       1, missing_data_2, error_format)
            Format_Excel_Cell(df, index, worksheet, blunder1, Youth_col,        1, missing_data_2, error_format)
            Format_Excel_Cell(df, index, worksheet, blunder1, Citizen_col,      1, missing_data_2, error_format)
            
    
    
    
    
    ## Format Unverified Training Interventions::
    for UTIID in unveri_TI_ID:
        Format_Excel_Cell(df, index, worksheet, UTIID, col_interven[1],         1, missing_data_2, unverified_TI_format)
        Format_Excel_Cell(df, index, worksheet, UTIID, col_type[1],             1, missing_data_2, unverified_TI_format)
        Format_Excel_Cell(df, index, worksheet, UTIID, col_NQF[1],              1, missing_data_2, unverified_TI_format)
        Format_Excel_Cell(df, index, worksheet, UTIID, col_duration[1],         1, missing_data_2, unverified_TI_format)
        Format_Excel_Cell(df, index, worksheet, UTIID, col_provider[1],         1, missing_data_2, unverified_TI_format)  
        

    ## Format residual columns 
    for RC1 in blunder_col_date_entered:
        Format_Excel_Cell(df, index, worksheet, RC1, col_date_entered[0],       1, missing_data_2, error_format)

    for RC2 in blunder_col_date_completed:
        Format_Excel_Cell(df, index, worksheet, RC2, col_date_completed[0],     1, missing_data_2, error_format)
    
    for RC3 in blunder_col_amount_spent:
        Format_Excel_Cell(df, index, worksheet, RC3, col_amount_spent[0],       1, missing_data_2, error_format)
    
    for RC4 in blunder_col_venue:
        Format_Excel_Cell(df, index, worksheet, RC4, col_venue[0],              1, missing_data_2, error_format)
    
    for RC5 in blunder_col_transport:
        Format_Excel_Cell(df, index, worksheet, RC5, col_transport[0],          1, missing_data_2, error_format)
    
    for RC6 in blunder_col_accommodation:
        Format_Excel_Cell(df, index, worksheet, RC6, col_accommodation[0],      1, missing_data_2, error_format)
        
    for RC7 in blunder_col_catering:
        Format_Excel_Cell(df, index, worksheet, RC7, col_catering[0],           1, missing_data_2, error_format)
        
    for RC8 in blunder_col_SETA_funded:
        Format_Excel_Cell(df, index, worksheet, RC8, col_SETA_funded[0],        1, missing_data_2, error_format)
        
    for RC9 in blunder_col_completed:
        Format_Excel_Cell(df, index, worksheet, RC9, col_completed[0],          1, missing_data_2, error_format)
        
    for RC10 in blunder_col_date_cert:
        Format_Excel_Cell(df, index, worksheet, RC10, col_date_cert[0],         1, missing_data_2, error_format)
        
    for RC11 in blunder_col_train_guide:
        Format_Excel_Cell(df, index, worksheet, RC11, col_train_guide[0],       1, missing_data_2, error_format)
        
    for RC12 in blunder_col_train_facil:
        Format_Excel_Cell(df, index, worksheet, RC12, col_train_facil[0],       1, missing_data_2, error_format)
        
    for RC13 in blunder_col_train_contact:
        Format_Excel_Cell(df, index, worksheet, RC13, col_train_contact[0],     1, missing_data_2, error_format)
        
    for RC14 in blunder_col_train_public:
        Format_Excel_Cell(df, index, worksheet, RC14, col_train_public[0],      1, missing_data_2, error_format)
        
    for RC15 in blunder_col_learner_local:
        Format_Excel_Cell(df, index, worksheet, RC15, col_learner_local[0],     1, missing_data_2, error_format)
        
    for RC16 in blunder_col_learner_urban:
        Format_Excel_Cell(df, index, worksheet, RC16, col_learner_urban[0],     1, missing_data_2, error_format)
        
    for RC17 in blunder_col_municipality:
        Format_Excel_Cell(df, index, worksheet, RC17, col_municipality[0],      1, missing_data_2, error_format)

    






    '''
    REVIST DUPLICATE PERSAL!!! - In event that there are still duplicate persals
    Broken code:
    '''    
#    ## Format Duplicate Persal #s
#    ## Redundant loop 
#    for index, row in df.iterrows():
#        if row[Dup_Check_1] in Dup_Persal:
#            try:
##                # row loc, column, cell:
##                print(df.loc[df[Dup_Check_1]==row[Dup_Check_1]].index[0],
##                      Dup_Check_1,
##                      row[Dup_Check_1])
#
#                worksheet.write(df.index.get_loc(index) +1,
#                                df.columns.get_loc(Dup_Check_1) +1, 
#                                row[Dup_Check_1],
#                                error_format)                
#            except:
#                pass
    
        
    writer.save()
    
    ## Possibly not an accurate measure of # of erroneous IDs
    ##  revisit !!! 
    no_err_ID = len(Blunder_ID) + len(bad_ID)
    
    no_unverify_ID = len(Unverified_ID)
    

    
    ##  (TEMP) Sheet Results 
    print('Sheet name: ',               sheet_name          )
    print('Duplicate rows:\t\t',        duplicate_rows      )
    print('Erroneous ID\'s:\t\t',       no_err_ID           )
    print('Unverified Persal:\t',       no_unverify_ID      )
    print('Age blunders:\t\t',          Glob_mis_age        )
    print('Gendr blunders:\t\t',        Glob_mis_gender     )
    print('Citiz blunders:\t\t',        Glob_mis_citizen    )
    print('Youth blunders:\t\t',        Glob_mis_youth      )
        
    print('Occ lvl blunders:\t',        Glob_mis_occlvl     )
    print('Job Title blunders:\t',      Glob_mis_jobtit     )
    print('Residence blunders:\t',      Glob_mis_res        )
    
    print(                                                  )
    print('First Name:\t\t',            Glob_mis_name       )
    print('Surname:\t\t',               Glob_mis_sname      )
    print('Persal No.:\t\t',            Glob_mis_persal     )
    print('Salary Level:\t\t',          Glob_mis_SLvL       )
    print('Race:\t\t\t',                Glob_mis_race       )
    print('Disability:\t\t',            Glob_mis_disab      )
    print(                                                  )
    print('Unverified T.I.:\t',         Glob_unveri_TI      )
    print('Intervention:\t\t',          Glob_intervention   )
    print('Type:\t\t\t',                Glob_TItype         )
    print('NQF:\t\t\t',                 Glob_NQF            )
    print('Duration:\t\t',              Glob_duration       )
    print('Provider:\t\t',              Glob_provider       )
    print(                                                  )
    print('Res. Issues:\t\t',           Glob_res_issues     )
    print(                                                  )
    
    
    Dirt_Report(parent_Dir, 
                folder_name_8,
                duplicate_rows,
                no_err_ID,
                no_unverify_ID,
                Glob_mis_age,
                Glob_mis_gender,
                Glob_mis_citizen,
                Glob_mis_youth,
                Glob_mis_name,
                Glob_mis_sname,
                Glob_mis_persal,
                Glob_mis_SLvL,
                Glob_mis_race,
                Glob_mis_disab,
                '-',
                'N/A',
                file_name.replace(file_type.replace('*',''),''),
                sheet_name,
                3,
                [],
                Glob_mis_occlvl,
                Glob_mis_jobtit,
                Glob_mis_res,
                Glob_unveri_TI,
                Glob_intervention,
                Glob_TItype,
                Glob_NQF,
                Glob_duration,
                Glob_provider,
                Glob_res_issues
                )


    
    
    return duplicate_rows, no_err_ID, no_unverify_ID, Glob_mis_age, Glob_mis_gender, Glob_mis_citizen, Glob_mis_youth, Glob_mis_name, Glob_mis_sname, Glob_mis_persal, Glob_mis_SLvL, Glob_mis_race, Glob_mis_disab, Glob_mis_occlvl, Glob_mis_jobtit, Glob_mis_res, Glob_unveri_TI, Glob_intervention, Glob_TItype, Glob_NQF, Glob_duration, Glob_provider, Glob_res_issues




def Critical_Error(err_reason, instruct_1):
    '''
    Feedback notification for critical errors
    '''
    
    'Error notifications'
    error_msg0  =   '[CRITICAL ERROR]'
    error_msg1  =   'Source of error:'
    error_msg2  =   'How to fix error:'
    error_msg3  =   '⯈ Replace missing file'
    error_msg4  =   '⯈ Redownload Data Cleanser'
    
    feedback.insert(INSERT, error_msg0 
                        + '\n\n' 
                        + error_msg1 + '\n' 
                        + err_reason
                        + '\n\n'
                        + error_msg2 + '\n' 
                        + error_msg3 
                        + instruct_1
                        + '\n\t(or)\n' 
                        + error_msg4)
        
    feedback.insert(END,'\n\n')  
    feedback.tag_add("description_4", "1.0", "1.17")
    feedback.tag_add("description_3", "3.0", "3.20")
    feedback.tag_add("description_3", "6.0", "6.20")
    feedback.configure(bg = '#ffc9cc')

def Critical_Error_2(err_reason):
    
    feedback.insert(INSERT, '[CRITICAL ERROR]'
                        + '\n\n' 
                        + 'Source of error:' + '\n' 
                        + err_reason)
    feedback.insert(END,'\n\n')  
    feedback.tag_add("description_4", "1.0", "1.17")
    feedback.tag_add("description_3", "3.0", "3.20")
    feedback.configure(bg = '#ffc9cc')
        
                       



def Other_Cols_Check_1(Glob_res_issues, row, col, index, blunder_list):
    '''
    Check for residual columns (columns that cannot be verified by databases) for valid monetary value 
    
        Checks for 'None' (or) float value
    '''

    try:
        if str(row[col]).lower() == 'none':
            pass
        elif isinstance(float(str(row[col]).lower().replace('r','').replace(' ','').replace(',','.')), float):
            pass
        else:
            blunder_list.append(index)
            Glob_res_issues += 1
    except ValueError:
        blunder_list.append(index)
        Glob_res_issues += 1
    except KeyError:
        ## Assumption: Column Header name variation not accounted for/ D.N.E
        print('KeyError: Other_Cols_Check_1')
        
            
    
    
    return blunder_list, Glob_res_issues 


def Other_Cols_Check_2(Glob_res_issues, row, index, col, blunder_list, Option_1, Option_2): 
    
    '''
    Check for residual columns for binary options

    '''
    try:
        string_seta = str(row[col]).lower().replace(' ','') 
        if string_seta == Option_1 or string_seta == Option_2:
            pass
        else:
            blunder_list.append(index)
            Glob_res_issues += 1
    except KeyError:
        ## Assumption: Column Header name variation not accounted for/ D.N.E
#        print('KeyError: Other_Cols_Check_2')
        pass
    
    return blunder_list, Glob_res_issues 



def Other_Cols_Check_3(Glob_res_issues, 
                       
                       df, index, row, 
                                              
                       col, blunder_list,
                       
                       string_1, string_2):
    
    '''
    Check for residual columns for valid date
    '''
    
    
    '''
    Method: 
        
    • Remove 'Common date seperators'           (slash [/], dash [-], period [.],comma [,]  )
    
    • Resultant value should be:
        • Eight digit number                    ( DDMMYYYY                                  )
        • Six digit number                      ( DDMMYY                                    )
        • Five digit number                     ( DMMYY                                     )
        • Month                                 ( Jan, March ...                            )
        • Day                                   ( Friday, Thur ...                          )
        • String                                ( 'Not', 'still waiting'                    )
         
    '''
    
    months_full     = ['January', 'February','March','April','May','June','July','August','September','October','November','December']
    days_full       = ['Monday','Tuesday','Wednesday','Thursday','Friday','Saturday','Sunday']
    month_slice     = []
    days_slice      = []
    for month in months_full:
        month_slice.append(month[0:3]) 
    for _day in days_full:
        days_slice.append(_day[0:3])
      
        
    
    
    try:
        
        cell0= str(row[col])
        cell = cell0
        cell = cell.replace('/','')
        cell = cell.replace('-','')
        cell = cell.replace('.','')
        cell = cell.replace(',','')
    
        cell = int(cell)
        cell = str(cell)
        
        if len(cell) == 8:
            pass
        
        elif len(cell) == 6:
            pass
        
        elif len(cell) == 5:
            ## append a 0 in front of DCI 
            df.at[index, col] = ('0' + cell0)
            Glob_res_issues
            
            
    except ValueError:
        str_check = 0
        
        ## Check for full month
        for fullmonth in months_full:
            if fullmonth.lower() in cell0.lower():
                str_check +=1
                break
            else:
                pass
        
        ## Check for truncated month 
        for truncmonth in month_slice:
            if truncmonth.lower() in cell0.lower():
                str_check +=1
                break
            else:
                pass
        
        for fullday in days_full:
            if fullday.lower() in cell0.lower():
                str_check +=1
                break
            else:
                pass
        
        for truncday in days_slice:
            if truncday.lower() in cell0.lower():
                str_check +=1
                break
            else:
                pass
        
        ## Check for other strings
        if string_1 in cell0.lower() or string_2 in cell0.lower():
            str_check +=1
        else:
            pass
        
        
        ## Add idices that did not contain any expected strings
        if str_check == 0:
            blunder_list.append(index)
            Glob_res_issues += 1
        else:
            pass
        
    except KeyError:
        ## Assumption: Column Header name variation not accounted for/ D.N.E
#        print('KeyError: Other_Cols_Check_3')
        pass
        
    return blunder_list, Glob_res_issues



def Format_Excel_Cell(df,index, worksheet,
                      element,
                      column, YY,
                      missing_text,
                      cell_format
                      ):
    
    '''
    Special formatting for uvnerified/ erroneous cells in output file
    '''
    
    
    '''
    Reminder!:
    
    
    df.iloc[X,Y] 
        returns cell:
            in row position (integer) X
            and col position (integer) Y
    
    
    df.index.get_loc(Z)
        returns integer row location of index 'Z'
        
    
     df.columns.get_loc('Column A')
         returns the integer position of 'Column A'
    
    
    
    df.iloc[  df.index.get_loc(float(blunder1))     ,    df.columns.get_loc(col_job1[1])  ]
    '''
    
    try:
        worksheet.write(df.index.get_loc(float(element)) + YY, df.columns.get_loc(column) + YY, 
                        df.iloc[df.index.get_loc(float(element)), df.columns.get_loc(column)], 
                        cell_format) 
    except TypeError:
        worksheet.write(df.index.get_loc(float(element)) + YY, df.columns.get_loc(column) + YY, 
                        missing_text, 
                        cell_format) 
    
    except KeyError:
        try:
            worksheet.write(df.index.get_loc(int(element)) + YY, df.columns.get_loc(column) + YY, 
                            df.iloc[df.index.get_loc(int(element)), df.columns.get_loc(column)], 
                            cell_format) 
        except KeyError:
            try:
                worksheet.write(df.index.get_loc(element) + YY, df.columns.get_loc(column) + YY, 
                                df.iloc[df.index.get_loc(element), df.columns.get_loc(column)], 
                                cell_format) 
            except TypeError:
                try:
                    worksheet.write(df.index.get_loc(element) + YY, df.columns.get_loc(column) + YY, 
                            missing_text, 
                            cell_format) 
                    
                except TypeError:
                    worksheet.write(df.index.get_loc(int(element)) + YY, df.columns.get_loc(column) + YY, 
                            missing_text, 
                            cell_format) 
#                    print('Type Error 2                         ', element )
            
            except KeyError:
                ## DNE or Unknown KeyError 
#                print('Key Error 3                              ', element )
                pass
            
    except ValueError:
        try:
            worksheet.write(df.index.get_loc(element) + YY, df.columns.get_loc(column) + YY, 
                                    df.iloc[df.index.get_loc(element), df.columns.get_loc(column)], 
                                    cell_format) 
        except KeyError:
            pass
        except TypeError:
            pass
        

    

def Training_Intervention_Verifier(index,           row,              TI_Index,        
                                   
                                   df,              df_db_TI,
                                   
                                   col_type,        Glob_TItype,
                                   col_NQF,         Glob_NQF,
                                   col_provider,    Glob_provider,
                                   col_duration,    Glob_duration,
                                   
                                   Verified_TI,     Unverified_TI,
                                                    unveri_TI_ID
                                   
                                   ):
    '''
    Verifier for Training Interventions:
        
        Weak- the intervention in QMR has to be identical to its counterpart in DB, else it will not be recognised
        
    '''

    try:
        
        db_TI_type      = df_db_TI.iloc[    df_db_TI.index.get_loc(  TI_Index  )][  col_type[0]         ]
        db_TI_NQF       = df_db_TI.iloc[    df_db_TI.index.get_loc(  TI_Index  )][  col_NQF[0]          ]
        db_TI_provider  = df_db_TI.iloc[    df_db_TI.index.get_loc(  TI_Index  )][  col_provider[0]     ]
        db_TI_duration1 = df_db_TI.iloc[    df_db_TI.index.get_loc(  TI_Index  )][  col_duration[0]     ]
        db_TI_duration2 = df_db_TI.iloc[    df_db_TI.index.get_loc(  TI_Index  )][  col_duration[2]     ]
        db_TI_duration  = str(db_TI_duration1) + ' ' + db_TI_duration2

        Verified_TI.append(TI_Index)
        
        if str(row[col_type[1]]).lower()        !=  str(db_TI_type).lower():
            df.at[index, col_type[1]]           =   db_TI_type
            Glob_TItype                         +=  1   
        else:
            pass
        
        if  str(row[col_NQF[1]]).lower()        !=  str(db_TI_NQF).lower():
            df.at[index, col_NQF[1]]            =   db_TI_NQF
            Glob_NQF                            +=  1  
        else:
            pass
                        
        if  str(row[col_provider[1]]).lower()   !=  str(db_TI_provider).lower():
            df.at[index, col_provider[1]]       =   db_TI_provider
            Glob_provider                       +=  1  
        else:
            pass
        
        if  str(row[col_duration[1]]).lower()   !=  db_TI_duration.lower():
            df.at[index, col_duration[1]]       =   db_TI_duration
            Glob_duration                       +=  1  
        else:
            pass
        
    except KeyError:
        ## Intervention D.N.E. in Training Intervention DB
        ## ∴ Cannot be verified 
        
        Unverified_TI.append(TI_Index)
        unveri_TI_ID.append(index)
        
        'ADD TO List for HIGHLIGHT'
    
    except AttributeError:
        print('AttributeError                       ',index)
        
#    except:
#        print('UnknownError                       ',index)
        
    return (Glob_TItype,
            Glob_NQF,
            Glob_provider,
            Glob_duration,
            
            Verified_TI,
            Unverified_TI,
            unveri_TI_ID)



'''
Needs fixing:
    Cater for all columns
'''
def Verify_Persal(df_db, df, row, index, ID_number, cols_verify):
    '''
    Incomplete function, not called
    Meant to verify Persal data
    A reduntant version of this method is used within Verify_Write (function)
    '''
    for CV in cols_verify:
        try:
            
            df_db_fname = sub(r"(\w)([A-Z])",r"\1 \2", (df_db.iloc[df_db.index.get_loc(int(ID_number))][CV[0]]).title().replace(' ',''))
            print(row[CV[1]],'\t\t', df_db_fname)
            if row[CV[1]]   !=  df_db_fname:
                
                df.at[index, CV[1]] = df_db_fname
            else:
                pass
        except:
            try:
                df_db_fname = sub(r"(\w)([A-Z])",r"\1 \2", (df_db.iloc[df_db.index.get_loc(int(ID_number))][CV[0]]).title().replace(' ',''))
                if sub(r"(\w)([A-Z])",r"\1 \2", row[CV[2]].title().replace(' ',''))   !=  df_db_fname:
                    df.at[index, CV[2]] = df_db_fname
                else:
                    pass
            except:
                #print(index,'D.N.E in ref df')
                pass
    #print('------------------------')
    return df
   



def DF_form(name, df, file_index_col, Dup_Check_3, Dup_Check_4):
    '''
    Set index & Drop Duplicates 
    '''
    
    '''
    Method 1:
    Drops all ID's with NaN values
    '''
#    ## select column as index
#    df = df.set_index(file_index_col) 
#    ## drop NaN values in index
#    ##  locate non null index values
#    df = df.loc[~df.index.isna()]  
    
    
    '''
    Method 2
    Check for rows with NaN ID's 
        > Drop if entire row is null
        > Keep if row has other data 
    '''
    
    '''
    Method 2 (EDITED)
    Cleans ID column and sets it as index for DF
    '''
    
    IDError = False
    
    try:
        missing = 1
        print(name)
        for index, row in df.iterrows():
            
             #print(row)
             ## Get ID number
             check_ID   = df.at[index, file_index_col]
             ## Get L.Name
             check_Sname= df.at[index, Dup_Check_3]
             ## Get Age
             check_age= df.at[index, Dup_Check_4]
             
              
             if isnull(check_ID) is True:
                 if isnull(check_Sname) is True:
                     ## If ID# and Sname is Null --> DROP
                     df = df.drop(float(index))
                 else:
                     if str.isdigit(str(check_age)) is True:
                         ## If age is an integer butID is missing:
                         ## Give arbitary value to missing ID
                         df.at[index, file_index_col] = 'Missing ID #'+str(missing)
                         #print(df.at[index, file_index_col])
                         missing += 1
                     else:
                        ## Age is also not a digit --> DROP
                        df = df.drop(index)
             elif isnull(check_ID) is False:
                 ## If a value for ID exists but it is not a digit --> DROP
                 if str.isdigit(str(df.at[index, file_index_col])) is False:
                     df = df.drop(index)
                 else:
                     pass
                     "Solution to IDs that are not 13 digits here???"
                     "Perhaps not -> deal with all erronous ID's together"
                     "instead of highlight red erroronous ID (digits) -> give it arbitrary name"
                    
        ## select column as index
        df = df.set_index(file_index_col)
        
        
        '''
        Drop duplicate rows 
        '''
        ## Rows with duplicate indices
        rows_all = df.shape[0] 
        ## drop duplicate indices 
        df = df.loc[~df.index.duplicated(keep='first')]
        ## Rows ~duplicate indices
        rows_after = df.shape[0]
        ## Number of duplicate rows
        dup_rows = rows_all - rows_after
    
    except KeyError:
        ##Index column header name DNE (or) differs from 'file_index_col'
        IDError = True
        dup_rows = 0
    
    
    return df, dup_rows, IDError
       


def Calc_Age(birthdate): 
    '''
    Calculate age using current date and D.O.B
    '''
    today = date.today() 
    try:  
        birthday = birthdate.replace(year = today.year) 
  
    except ValueError:  
        birthday = birthdate.replace(year = today.year, month = birthdate.month + 1, day = 1) 
        
    if birthday > today: 
        return today.year - birthdate.year - 1
    else: 
        return today.year - birthdate.year 



# =============================================================================
# =============================================================================
# =============================================================================
# =============================================================================
# =============================================================================

    
    
    
    
    
    


    








# ~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~ 
# ~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~ 
# ~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~ 
# ~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~ 
# ~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~ 
# ~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~ 
# ~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~ 
# ~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~ 
# ~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~
# ~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~
# ~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~
# ~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~
# ~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~
# ~~~~~~~~~~~~~~~~~~~~~~~~~~~~      GUI     ~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~
# ~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~
# ~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~
# ~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~
# ~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~
# ~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~ 
# ~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~ 
# ~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~ 
# ~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~ 
# ~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~ 
# ~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~ 
# ~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~ 
# ~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~ 
# ~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~ 




"Functions called from GUI"
def About():
    ## message box
    messagebox.showinfo('About', Texts.Description())
def Version():
    messagebox.showinfo('Version', Texts.Version(version_string, last_updated))
def Report():
    messagebox.showinfo('Report Issues', Texts.Report())


def About_2(): 
    ## Slave window
    Slave4 = Toplevel()
    Slave4.title('About')
    Slave4.geometry("300x500")
    Slave4.configure(background='white')
    Slave4.resizable(False, False)
    Slave4.iconbitmap(image_dir+'Icon0.ico')
    
    text_about  = Texts.Description()
    try:
        photo_SD    = PhotoImage(file = getcwd()+'\\' + 'Image_files\\' + 'IM_logo.png') 
    except TclError:
        pass
    
    
    Frame_about = Frame(Slave4, background = 'white')
    Frame_logo2 = Frame(Slave4, background ='white')
    
    
    Text_about = Text(Frame_about, 
               bg ='#FFFFFF', 
               fg ='black', 
               highlightbackground='#FFFFFF', 
               bd = 0, cursor ="arrow", 
               font = 'Open_Sans 11', 
               height = 23, 
               padx = 0, pady = 0,
               width = 25, 
               wrap = CHAR)
    
    Text_about.insert(INSERT, text_about)
    Text_about.insert(END,"")
    Text_about.tag_configure("left", justify='left')
    Text_about.tag_add("left", 1.0, "end")
    Text_about.configure(state="disabled")
    Text_about.grid()#row=0, column =0, sticky=W, padx =0, pady=0)
        
    

    try:
        Label_logo2 = Label(Frame_logo2, width=272, height=91, bg='white', image=photo_SD)
        Label_logo2.grid()#row=0, column =0, sticky=W, padx =50, pady=0)
    except UnboundLocalError:
        pass
    
    
    
    Frame_about.pack()#fill=X)
    Frame_logo2.pack()#fill=X)
    
    
    
    mainloop()
    
    

#    Title_1 = Text(Frame_Title_1, 
#                   bg ='#FFFFFF', 
#                   fg ='#02319A', 
#                   highlightbackground='#FFFFFF', 
#                   bd = 0, cursor ="arrow", 
#                   font = 'Open_Sans 25 bold', 
#                   height = 1, 
#                   padx = 0, pady = 0,
#                   width = 25, 
#                   wrap = CHAR)
#    
#    Title_1.insert(INSERT, string_title_1)
#    Title_1.insert(END,"")
#    Title_1.tag_configure("center", justify='center')
#    Title_1.tag_add("center", 1.0, "end")
#    Title_1.configure(state="disabled")
#    Title_1.grid(row=0, column =0, sticky=W, padx =0, pady=0)
#        
#    
    
    
    
    
    
    
    
    
def Changelog(): 
    Slave5 = Toplevel()
    Slave5.title('Changelog')
    Slave5.geometry("800x520")
    Slave5.configure(background='white')
    Slave5.resizable(False, False)
    Slave5.iconbitmap(image_dir+'Icon0.ico')
    
    text_changes  = Texts.Changelog_text()
    photo_SD    = PhotoImage(file = getcwd()+'\\' + 'Image_files\\' + 'IM_logo.png') 
    
    
    
    Frame_changes = Frame(Slave5, background = 'white')
    Frame_logo2   = Frame(Slave5, background ='white')
    ybar= Scrollbar(Frame_changes)
    
    
    
    Text_about = Text(Frame_changes, 
                   bg ='#FFFFFF', 
                   fg ='black', 
                   highlightbackground='#FFFFFF', 
                   bd = 1, cursor ="arrow", 
                   font = 'Helvetica 11', 
                   height = 25, 
                   padx = 0, pady = 0,
                   width = 97, 
                   wrap = CHAR)
    
    Text_about.insert(INSERT, text_changes)
    Text_about.insert(END,"")
    Text_about.tag_configure("left", justify='left')
    Text_about.tag_add("left", 1.0, "end")
    Text_about.configure(state="disabled")
    
    ybar.config(command=Text_about.yview)
    Text_about.config(yscrollcommand=ybar.set)
    
    Text_about.grid()
    ybar.grid(row=0, column=1, sticky="ns")
    
    
    

    Label_logo2 = Label(Frame_logo2, width=272, height=91, bg='white', image=photo_SD)
    Label_logo2.grid()#row=0, column =0, sticky=W, padx =50, pady=0)
    
    
    Frame_changes.pack()#fill=X)
    Frame_logo2.pack()#fill=X)
    
    
    
    mainloop()
        
    
    
    
    
    
 
    
def Browse_Dir():
    '''
    Window to browse for excel file to process
    '''
    try:
        Browse_str = filedialog.askopenfilename(filetypes=[("Excel files","*.xlsx .xls") , ("All files", "*")])                                   
        s_file = Browse_str.split('/')[-1]
        Copy_Over(Browse_str, s_file)
    except:
        print('No file Selected')
        feedback.delete('1.0', END)
        feedback.insert(INSERT, '\n' + '⯈ No file Selected')
        feedback.insert(END, '\n\n')    


def Copy_Over(Browse_str, s_file):
    '''
    Copy over file selected in Browse_Dir
    '''
    parent_Dir  =  getcwd() 
    tar_dir     = parent_Dir + '\Input files' 
    Copy_file(Browse_str, tar_dir)
    feedback.delete('1.0', END)
    feedback.insert(INSERT, '\n' + '⯈ File copied:\t\t' + s_file)
    feedback.insert(END, '\n\n')     
        
                                                                    
def Output_Format(root_x, root_y, dx, dy):
    '''
    Slave window for xlsx output options
    '''                                                            
    Slave2 = Toplevel()
    Slave2.title('Output Format')
    Slave2.geometry("200x230") 
    Slave2.configure(background='#02319A')
    Slave2.resizable(False, False)
    Slave2.iconbitmap(image_dir+'settings_icon.ico')
    
    Slave2.focus_set()                                                        
    Slave2.grab_set()   
    
    
    Slave2_x = Slave2.winfo_width()
    Slave2_y = Slave2.winfo_height()  
    
    Slave2.geometry("%dx%d+%d+%d" % (Slave2_x, Slave2_y, root_x + dx, root_y + dy))
    
    
    Slave2_Frame_A = Frame(Slave2, background='#5C7EF9' )
    Slave2_Frame_B = Frame(Slave2, background='#5C7EF9' )  
    Slave2_Frame_C = Frame(Slave2, background='#5C7EF9' )  

    output_choice = StringVar()
    output_choice.set("Separate Sheets")
    
    Options = [
            ("Formatted Workbook",      "Formatted Workbook"    ),
            ("Simple Workbook",         "Simple Workbook"       ),
            ("Separate Sheets",         "Separate Sheets"       )
            ]
    
    def Clicked():
        '''

        '''
        global radio_outformat
        radio_outformat = output_choice.get()
        
        feedback.delete('1.0', END)
        feedback.insert(INSERT, '\n' + '⯈ Output Format Selected:\t\t' + radio_outformat)
        feedback.insert(END, '\n\n') 

    
    for text, value in Options:
        Radiobutton(Slave2_Frame_B, 
                    bg = '#5C7EF9',
                    fg = 'black',
                    text = text, 
                    variable = output_choice,   
                    command = Clicked,
                    value = value).pack(anchor=W)
        
        
    Instruct1 = Text(Slave2_Frame_A, 
               bg ='#5C7EF9', fg ='#ffffff', 
               highlightbackground='#FFFFFF', 
               bd = 0, 
               cursor ="arrow", 
               font = 'Helvetica 12 bold', 
               height = 4, 
               padx = 0, pady = 0,
               width = 21, 
               wrap = CHAR)
    Instruction_1 = '\nSelect format for\noutput workbook:'
    Instruct1.insert(INSERT, Instruction_1)
    Instruct1.insert(END,"")
    Instruct1.tag_configure("center", justify='center')
    Instruct1.tag_add("center", 1.0, "end")
    Instruct1.configure(state="disabled")
    Instruct1.grid(row=0, column =0, sticky=W, padx =0, pady=0)
    
    
    Instruct2 = Text(Slave2_Frame_C, 
               bg ='#5C7EF9', fg ='#ffffff', 
               highlightbackground='#FFFFFF', 
               bd = 0, 
               cursor ="arrow", 
               font = 'Helvetica 9', 
               height = 4, 
               padx = 0, pady = 0,
               width = 28, 
               wrap = CHAR)
    Instruction_2 = '\nFor help, please visit:\nHelp ⯈ User Guide'
    Instruct2.insert(INSERT, Instruction_2)
    Instruct2.insert(END,"")
    Instruct2.tag_configure("center", justify='center')
    Instruct2.tag_add("center", 1.0, "end")
    Instruct2.configure(state="disabled")
    Instruct2.grid(row=0, column =0, sticky=W, padx =0, pady=10)
        
    Slave2_Frame_A.pack(fill=X) 
    Slave2_Frame_B.pack(fill=X) 
    Slave2_Frame_C.pack(fill=X) 


def Graphics_Window(root_x, root_y, dx, dy):
    '''
    Window for graphics options
    ''' 
    ## Slave window
    Slave3 = Toplevel()
    Slave3.title('Graphics')
    Slave3.geometry("200x230")
    Slave3.configure(background='#5C7EF9')
    Slave3.resizable(False, False)
    Slave3.iconbitmap(image_dir+'graphics_icon.ico')
    
    #Slave3.focus_set()                                                        
    Slave3.grab_set()  
    
    Slave3_x = Slave3.winfo_width()
    Slave3_y = Slave3.winfo_height()  
    
    Slave3.geometry("%dx%d+%d+%d" % (Slave3_x, Slave3_y, root_x + dx, root_y + dy))
    
    ## Radio buttons for Graph type
    graphic_choice = StringVar()
    graphic_choice.set("Bar Graph")
    
    
    Options2 = [("Bar Graph",   "Bar Graph")]
#    Options2 = [
#        ("Bar Graph",   "Bar Graph"),
#        ("Pie Chart",   "Pie Chart")
#        ]
    
        
    
    def Clicked2():
        '''
        Get User choice for graphics type
        '''
        global radio_graphic
        radio_graphic = graphic_choice.get()
        #print(radio_graphic)
        
        feedback.delete('1.0', END)
        feedback.insert(INSERT, '\n' + '⯈ Graphic output selected:\t\t' + radio_graphic)
        feedback.insert(END, '\n\n') 
    
    
    Label_G1= Label(Slave3, 
                    width=30, height=1, 
                    bg='#5C7EF9', fg= 'white', 
                    text ='Output Figure:', font = 'Open_Sans 12 bold',
                    anchor=W)
    Label_G1.pack(anchor=W)
    
    for text, value in Options2:
        Radiobutton(Slave3, 
                    bg = '#5C7EF9',
                    fg = 'black',
                    text = text, 
                    variable = graphic_choice,   
                    command = Clicked2,
                    value = value).pack(anchor=W)
    
    
    ## Radio buttons for state of Graph output (Active/Not active)
    graphic_choice2 = StringVar()
    graphic_choice2.set("Not Active")
    
    Options3 = [
        ("Activate",    "Active"),
        ("Deactivate",  "Not Active")
        ]
    
    def Clicked3():
        '''
        Get User choice for graphic state
        '''
        global radio_graphic2
        radio_graphic2 = graphic_choice2.get()
        #print(radio_graphic2)
        
        feedback.delete('1.0', END)
        feedback.insert(INSERT, '\n' + '⯈ Graphic output state:\t\t' + radio_graphic2)
        feedback.insert(END, '\n\n') 
        
    Label_G2= Label(Slave3, 
                    width=30, height=1, 
                    bg='#5C7EF9', fg= 'white', 
                    text ='Figure output State:', font = 'Open_Sans 12 bold',
                    anchor=W)
    Label_G2.pack(anchor=W)
        
    for text, value in Options3:
        Radiobutton(Slave3, 
                    bg = '#5C7EF9',
                    fg = 'black',
                    text = text, 
                    variable = graphic_choice2,   
                    command = Clicked3,
                    value = value).pack(anchor=W)
        
    Label_G3= Label(Slave3, 
                width=30, height=1, 
                bg='#5C7EF9', fg= 'white', 
                text ='', font = 'Open_Sans 12 bold',
                anchor=W)
    Label_G3.pack(anchor=W)
    
    
    ## Open Directory containing output graphs 
    button_graphdir = Button(Slave3, 
                    text = "Open location of Graphics", 
                    font = 'Open_Sans 9', 
                    bg ='#5C7EF9', fg='White',
                    command = lambda:  startfile(  getcwd() + '\\''' + 'Result Diagrams'))

    button_graphdir.config( height = 2, width = 30 )
    button_graphdir.pack(anchor=W)


def Open_Dir(file_directory):
     startfile(  getcwd() + '\\''' + file_directory)
    
def Cleansing_Complete(counter_book):
    if counter_book == 1:
        book = ' book'
    else:
        book = ' books'
        
    messagebox.showinfo('Cleansing Complete','Successfully cleansed '
                        + str(counter_book)
                        + book)

def Open_file(filename):
    '''
    Open file from provided dir. 
    '''
    try:
        startfile(filename)
    except FileNotFoundError:
        messagebox.showinfo('Error', 'File is missing!\nYou probably deleted it :/')

def Fatal_Error():
    messagebox.showinfo('FATAL ERROR!', 'Data Cleanser failed to start properly!\nCritical files are missing')



'Incomplete E.Egg'
#def Easter_Egg():
#    SlaveEgg = Toplevel()
#    SlaveEgg.title('Trivia')
#    SlaveEgg.geometry("1000x1000")
#    SlaveEgg.configure(background='white')
#    SlaveEgg.resizable(False, False)
#    SlaveEgg.iconbitmap(image_dir+'Icon0.ico')
#    
# 
#    Frame_Ascii     = Frame(SlaveEgg, background = 'white')
#    Frame_Question  = Frame(SlaveEgg, background = 'white')
#    
#    
#    Text_Ascii = Text(Frame_Ascii, 
#                   bg ='#FFFFFF', 
#                   fg ='#02319A', 
#                   highlightbackground='#FFFFFF', 
#                   bd = 0, cursor ="arrow", 
#                   font = 'Open_Sans 10', 
#                   height = 300, width = 200, 
#                   padx = 0, pady = 0,
#                   wrap = CHAR)
#    
#    Ascii  = Texts.Ascii_Art()
#    Text_Ascii.insert(INSERT, Ascii)
#    Text_Ascii.insert(END,"")
#    Text_Ascii.tag_configure("left", justify='left')
#    Text_Ascii.tag_add("left", 1.0, "end")
#    Text_Ascii.configure(state="disabled")
#    Text_Ascii.grid()
#        
#
#    
#    Frame_Ascii.pack()#fill=X)
#    Frame_Question.pack()#fill=X)
#    
#    mainloop()
    

"Hover Button Highlights"
#def on_enter1(e):
#    button_graphics['bg'] = '#F82C81'
#
#def on_leave1(e):
#    button_graphics['bg'] = '#FFFFFF'
#
#def on_enter2(e):
#    button_outformat['bg'] = '#B428C7'
#def on_leave2(e):
#    button_outformat['bg'] = '#FFFFFF'
#
#def on_enter3(e):
#    button_browse['bg'] = '#5C7EF9'
#def on_leave3(e):
#    button_browse['bg'] = '#FFFFFF'
#    
#def on_enter4(e):
#    button['bg'] = '#168FFA'
#def on_leave4(e):
#    button['bg'] = '#FFFFFF'
#
#


#def on_enter1(e):
#    button_graphics['bg'] = '#5C7EF9'
#
#def on_leave1(e):
#    button_graphics['bg'] = '#FFFFFF'
#
#def on_enter2(e):
#    button_outformat['bg'] = '#5C7EF9'
#def on_leave2(e):
#    button_outformat['bg'] = '#FFFFFF'
#
#def on_enter3(e):
#    button_browse['bg'] = '#5C7EF9'
#def on_leave3(e):
#    button_browse['bg'] = '#FFFFFF'
#    
#def on_enter4(e):
#    button['bg'] = '#5C7EF9'
#def on_leave4(e):
#    button['bg'] = '#FFFFFF'
#
#def on_enter5(e):
#    button_inputdir['bg'] = '#5C7EF9'
#def on_leave5(e):
#    button_inputdir['bg'] = '#FFFFFF'
#    
#def on_enter6(e):
#    button_outputdir['bg'] = '#5C7EF9'
#def on_leave6(e):
#    button_outputdir['bg'] = '#FFFFFF'
    
    


def on_enter1(e):
    button_graphics['bg'] = '#c4d1ff'

def on_leave1(e):
    button_graphics['bg'] = '#FFFFFF'

def on_enter2(e):
    button_outformat['bg'] = '#c4d1ff'
def on_leave2(e):
    button_outformat['bg'] = '#FFFFFF'

def on_enter3(e):
    button_browse['bg'] = '#c4d1ff'
def on_leave3(e):
    button_browse['bg'] = '#FFFFFF'
    
def on_enter4(e):
    button['bg'] = '#c4d1ff'
def on_leave4(e):
    button['bg'] = '#FFFFFF'

def on_enter5(e):
    button_inputdir['bg'] = '#c4d1ff'
def on_leave5(e):
    button_inputdir['bg'] = '#FFFFFF'
    
def on_enter6(e):
    button_outputdir['bg'] = '#c4d1ff'
def on_leave6(e):
    button_outputdir['bg'] = '#FFFFFF'



"Tooltip"
#class ToolTip(object):
#
#    def __init__(self, widget):
#        self.widget = widget
#        self.tipwindow = None
#        self.id = None
#        self.x = self.y = 0
#
#    def showtip(self, text):
#        "Display text in tooltip window"
#        self.text = text
#        if self.tipwindow or not self.text:
#            return
#        x, y, cx, cy = self.widget.bbox("insert")
#        x = x + self.widget.winfo_rootx() + 57
#        y = y + cy + self.widget.winfo_rooty() +27
#        self.tipwindow = tw = Toplevel(self.widget)
#        tw.wm_overrideredirect(1)
#        tw.wm_geometry("+%d+%d" % (x, y))
#        label = Label(tw, text=self.text, justify=LEFT,
#                      background="#ffffff", relief=SOLID, borderwidth=1,
#                      font=("tahoma", "12", "normal"))
#        label.pack(ipadx=1)
#
#    def hidetip(self):
#        tw = self.tipwindow
#        self.tipwindow = None
#        if tw:
#            tw.destroy()
#
#def CreateToolTip(widget, text):
#    toolTip = ToolTip(widget)
#    def enter(event):
#        toolTip.showtip(text)
#    def leave(event):
#        toolTip.hidetip()
#    widget.bind('<Enter>', enter)
#    widget.bind('<Leave>', leave)
#    




def DNP():
    pass

# ~~~~~~~~~~~~~~~~~~~~~~~~~~~~ Master Window ~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~

image_dir =  getcwd() + '\\''' + 'Image_files\\'

root = Tk()
root.geometry("440x750")
root.title('Data Cleanser v'+version_string)
root.configure(background='white')
try:
    root.iconbitmap(image_dir+'Icon0.ico')
except TclError:
    Fatal_Error()
root.resizable(False, False)

## coordinates of root window
root_x = root.winfo_x()
root_y = root.winfo_y()

## window offesets 
dx = 100
dy = 300

## User's screen size
win_w, win_h = root.winfo_screenwidth(), root.winfo_screenheight()


# ~~~~~~~~~~~~~~~~~~~~~~~~~~~~ Dropdown Menus ~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~

menu = Menu(root)
root.config(menu=menu)


## Help Menu
dd_Help = Menu(menu)
menu.add_cascade(label="Help",menu=dd_Help)
dd_Help.add_command(label="About",command=About_2)
dd_Help.add_command(label="Version",command=Version)
dd_Help.add_command(label="User Guide", 
    command = lambda: Open_file(getcwd()+'\\''' + 'Documents\\' + 'User_Guide.pdf'))                          
dd_Help.add_separator()   
      
#dd_Help.add_command(label="Changelog",
#    command = lambda: Open_file( getcwd()+'\\''' + 'Documents\\' + 'Changelog.txt'))
dd_Help.add_command(label="Changelog", 
                    command = lambda: Changelog())  

dd_Help.add_command(label="Release Notes", 
    command = lambda: Open_file( getcwd()+'\\''' + 'Documents\\' + 'Relase_notes.pdf'))  




dd_Help.add_separator()
dd_Help.add_command(label="Report Issues",command=Report)                       
                                    
## Settings
dd_Settings = Menu(menu)
menu.add_cascade(label="Baseline Data",menu=dd_Settings)
dd_Settings.add_command(label="Open Baseline Data directory", command=lambda:Open_Dir('DB')) # TEMP? Command

## Other
dd_Other = Menu(menu)
menu.add_cascade(label="Reports",menu=dd_Other)
dd_Other.add_command(label="Results logs", command=lambda:Open_Dir('Result Reports'))   
#dd_Other.add_command (label="E.Egg", command=lambda:Easter_Egg())                            



# ~~~~~~~~~~~~~~~~~~~~~~~~~~~~  Frame() ~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~

Frame_logo          = Frame(root, background='white' )
Frame_banner        = Frame(root, background='white' )
Frame_button        = Frame(root, background='#FFFFFF' )
Frame_button_browse = Frame(root, background='#FFFFFF' )
Frame_Title_1       = Frame(root, background='#FFFFFF' )
Frame_Title_2       = Frame(root, background='#FFFFFF' )
Frame_descrip       = Frame(root, background='#FFFFFF' ,
                            padx = 15,)
#Frame_gap           = Frame(root, background='#FFFFFF' ) 

Frame_feedback      = Frame(root, background='#FFFFFF' ,
                            padx = 15, pady = 10) 



                        
                        
                        
# ~~~~~~~~~~~~~~~~~~~~~~~~~~~~  Label() ~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~

try:
    ## Label for logo
    photo = PhotoImage(file = image_dir + 'grad21.png' )
    Label_logo = Label(Frame_logo, width=446, height=160, bg='white', image=photo)
    Label_logo.grid(row=0, column =0, sticky=W, padx =0, pady=0)
    
    #label_gap = Label(Frame_logo, width=0, height=0, bg='white', image=photo)
    
    
    ## Label for banner
    photo_banner = PhotoImage(file = image_dir + 'banner1.png' )
    Label_banner = Label(Frame_banner, width=500, height=10, bg='white', image=photo_banner)
    Label_banner.grid(row=0, column =0, sticky=W, padx =0, pady=0)
except TclError:
    pass


# ~~~~~~~~~~~~~~~~~~~~~~~~~~~~  Text() ~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~
'''
Explode the below !!! 
'''

string_title_1  = " Data Cleanser"
string_title_2  = ""
string_title_2B = "for Quarterly Skills Development Reports"

# Title 1 
Title_1 = Text(Frame_Title_1, 
               bg ='#FFFFFF', fg ='#02319A', 
               highlightbackground='#FFFFFF', 
               bd = 0, 
               cursor ="arrow", 
               font = 'Helvetica 45 bold', 
               height = 1, 
               padx = 0, pady = 0,
               width = 25, 
               wrap = CHAR)

Title_1.insert(INSERT, string_title_1)
Title_1.insert(END,"")
Title_1.tag_configure("left", justify='left')
Title_1.tag_add("center", 1.0, "end")
Title_1.configure(state="disabled")
Title_1.grid(row=0, column =0, sticky=W, padx =0, pady=0)

# Title 2  
Title_2 = Text(Frame_Title_2, 
               bg ='#FFFFFF', fg ='#000000', 
               highlightbackground='#FFFFFF', 
               bd = 0, 
               cursor ="arrow", 
               font = 'Helvetica 10 bold', 
               height = 0, width = 63,
               padx = 0, pady = 0,
               wrap = CHAR
               )


Title_2.insert(INSERT, string_title_2B)
Title_2.insert(END,"")
Title_2.tag_configure("center", justify='center')
Title_2.tag_add("center", 1.0, "end")
Title_2.configure(state="disabled")
Title_2.grid(row=0, column =0, sticky=W, padx =0, pady=0)



string_descrip_1 = "\nDescription:"
string_descrip_2 = '\n• Verifies all input data'
string_descrip_3 = '\n• Corrects data where applicable'
string_descrip_4 = '\n• Where data cannot be fixed, it will be highlighted in output file'
string_descrip_5 = '\n\n'
string_descrip_6 = 'Instructions:'
string_descrip_7 = '\n• Use "Find file" to find your file'
string_descrip_8 = '\n• Click "Cleanse now" to commence data cleansing'

string_descrip_9 = '\n• Checks and Corrects all data'



## "Description Box"
Descrip = Text(Frame_descrip, 
               bg ='#ffffff', fg ='#000000', 
               highlightbackground='#FFFFFF',
               bd = 0, cursor ="arrow", 
               font = 'Helvetica 9', 
               height = 1, width = 57, 
               padx = 10, pady = 10,
               wrap = CHAR) 

#Descrip = Text(Frame_descrip, 
#               bg ='#f7fcff', fg ='#000000', 
#               highlightbackground='#FFFFFF',
#               bd = 1, cursor ="arrow", 
#               font = 'Helvetica 9', 
#               height = 8, width = 57, 
#               padx = 10, pady = 10,
#               wrap = CHAR) 
#
#Descrip.insert(INSERT, 
#               string_descrip_1 
#               + string_descrip_9 
#               + string_descrip_4 
#               + string_descrip_5
#               )
#
#Descrip.insert(INSERT, string_descrip_6 + string_descrip_7 +string_descrip_8)
#Descrip.tag_add("description_1", "1.0", "1.12")
#Descrip.tag_config("description_1", foreground="#02319A")#,  underline=1)
#Descrip.tag_add("description_2", "5.0", "5.13")
#Descrip.tag_config("description_2", foreground="#02319A")#,  underline=1)
#Descrip.configure(state="disabled")
#Descrip.insert(END,"")
Descrip.grid(row=0, column =0, sticky=W, padx =0, pady=0)


# "Feedback box"


ybar= Scrollbar(Frame_feedback)

feedback = Text(Frame_feedback, 
                fg ='#000000', bg ='#ffffff', 
                bd = 0, 
                cursor ="arrow", 
                font = "Helvetica 9",  
                padx = 10, pady = 0,
                height = 10, width = 55, 
                wrap = CHAR)


feedback.insert(INSERT, 
               string_descrip_1 
               + string_descrip_9 
               + string_descrip_4 
               + string_descrip_5
               )

feedback.insert(INSERT, string_descrip_6 + string_descrip_7 +string_descrip_8)
feedback.tag_add("description_1", "2.0", "2.12")
feedback.tag_config("description_1", foreground="#168FFA")#,  underline=1)
feedback.tag_add("description_2", "6.0", "6.13")
feedback.tag_config("description_2", foreground="#168FFA")#,  underline=1)
feedback.tag_config("description_3", foreground="#e80021", underline=1)
feedback.tag_config("description_4", foreground="#e80021")
#feedback.configure(state="disabled")
feedback.insert(END,"")
#feedback.grid(row=0, column =0, sticky=W, padx =0, pady=0)


#feedback.insert(INSERT, '\nThis is the Feedback Window...\n\n\nPlease see the \'User Guide\' under the \'Help\' menu for detailed instructions ')
##feedback.configure(state="disabled")
#feedback.insert(END,"")

ybar.config(command=feedback.yview)
feedback.config(yscrollcommand=ybar.set)


feedback.grid(row=0, column =0, sticky=W, padx =0, pady=0)
ybar.grid(row=0, column=1, sticky="ns")







# ~~~~~~~~~~~~~~~~~~~~~~~~~~~~  Button() ~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~
try:
    try:
        photo2 = PhotoImage(file = image_dir + 'sheet.png') 
        photo2 = photo2.subsample(8, 8) 
        
        photo3 = PhotoImage(file = image_dir + 'settings.png') 
        photo3 = photo3.subsample(1, 1) 
        
        photo4 = PhotoImage(file = image_dir + 'graphics.png') 
        photo4 = photo4.subsample(1, 1) 
        
        photo5 = PhotoImage(file = image_dir + 'Clean_1.png') 
        photo5 = photo5.subsample(6, 6) 
        
        photo6 = PhotoImage(file = image_dir + 'browse_files.png') 
        photo6 = photo6.subsample(1, 1) 
    except TclError:
        pass
    
    
    
    " Graphics "
    button_graphics = Button(Frame_button, 
                           text = " Graphics", 
                           font = 'Helvetica 13', 
                           bg ='white', fg='#02319A', 
                           image = photo4, compound = LEFT,
                           height = 60, width = 200, 
                           borderwidth=0,
                           command = lambda: Graphics_Window(root_x, root_y, dx, dy))
    button_graphics.grid(row=0, column =0, 
                         padx =10, pady=13)
    button_graphics.bind("<Enter>", on_enter1)
    button_graphics.bind("<Leave>", on_leave1)
    #CreateToolTip(button_graphics, text = 'Graphics')
    
    
    
    " Output format "
    button_outformat = Button(Frame_button, 
                           text = " Output Format", 
                           font = 'Helvetica 13', 
                           bg ='white', fg='#02319A', 
                           image = photo3, compound = LEFT,
                           height = 60, width = 200, 
                           borderwidth=0,
                           command = lambda: Output_Format(root_x, root_y, dx, dy))
    button_outformat.grid(row=0, column =1, 
                          padx =0, pady=13)
    button_outformat.bind("<Enter>", on_enter2)
    button_outformat.bind("<Leave>", on_leave2)
    #CreateToolTip(button_outformat, text = 'Output format')
    
    
    
    " Find file "
    button_browse = Button(Frame_button, 
                           text = "   Find file", 
                           font = 'Helvetica 13', 
                           bg ='white', fg='#02319A', 
                           image = photo2, compound = LEFT,
                           height = 60, width = 200, 
                           borderwidth=0,
                           command = lambda: Browse_Dir())
    #button_browse.config( height = 2, width = 15 )
    button_browse.grid(row=2, column =0, 
                       padx =10, pady=13)
    button_browse.bind("<Enter>", on_enter3)
    button_browse.bind("<Leave>", on_leave3)
    #CreateToolTip(button_browse, text = 'Find file')
    
    
    radio_outformat = 'Separate Sheets'
    radio_graphic  = 'Bar Graph'
    radio_graphic2 = 'Not Active'
    
    
    ## Cleanse Button v1
    #button = Button(Frame_button, 
    #                text = "Cleanse now", 
    #                font = 'Open_Sans 11', 
    #                bg ='#02319A', fg='White',
    #                command = lambda: main(radio_outformat, radio_graphic, radio_graphic2))
    #button.config( height = 2, width = 15 )
    #button.grid(row=1, column =1, padx =0, pady=5)
    
    
    
    " Cleanse Button v2 "
    button = Button(Frame_button, 
                   text = "Cleanse Data", 
                   font = 'Helvetica 13 underline', 
                   fg ='#02319A', bg='White', 
                   image = photo5, compound = LEFT,
                   height = 60, width = 200, 
                   borderwidth=0,
                   command = lambda:[DNP(), main(radio_outformat, radio_graphic, radio_graphic2, win_w, win_h)])
    
    
    button.grid(row=2, column =1, 
                padx =0, pady=13)
    button.bind("<Enter>", on_enter4)
    button.bind("<Leave>", on_leave4)
    #CreateToolTip(button, text = 'Cleanse Data')
    
    
    
    " Input dir "
    button_inputdir = Button(Frame_button, 
                           text = " Input files", 
                           font = 'Helvetica 13', 
                           bg ='white', fg='#02319A', 
                           image = photo6, compound = LEFT,
                           height = 60, width = 200, 
                           borderwidth=0,
                           command = lambda: startfile(  getcwd() + '\\''' + 'Input files'))
    button_inputdir.grid(row=1, column =0, 
                         padx =10, pady=13)
    button_inputdir.bind("<Enter>", on_enter5)
    button_inputdir.bind("<Leave>", on_leave5)
    #CreateToolTip(button_inputdir, text = 'Input files')
    
    
    
    " Output dir "
    button_outputdir = Button(Frame_button, 
                           text = "   Output files", 
                           font = 'Helvetica 13', 
                           bg ='white', fg='#02319A', 
                           image = photo6, compound = LEFT,
                           height = 60, width = 200, 
                           borderwidth=0,
                           command = lambda: startfile(  getcwd() + '\\''' + 'Output files'))
    button_outputdir.grid(row=1, column =1, 
                         padx =0, pady=13)
    button_outputdir.bind("<Enter>", on_enter6)
    button_outputdir.bind("<Leave>", on_leave6)
    #CreateToolTip(button_outputdir, text = 'Output files')
    
    
    
    #Frame_Button_browse
except NameError:
    pass





# ~~~~~~~~~~~~~~~~~~~~~~~~~~~~  pack() ~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~

 
Frame_logo.pack(fill=X)  
Frame_Title_1.pack(fill=X)
Frame_Title_2.pack(fill=X)
Frame_descrip.pack(fill=X) 
#Frame_gap.pack(fill=X)   
Frame_button.pack(fill=X) 
#Frame_button_browse.pack(fill=X)
Frame_feedback.pack(fill=X)  
Frame_banner.pack(fill=X) 



                 
root.mainloop()

